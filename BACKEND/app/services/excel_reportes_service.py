"""
Export a Excel (XLSX) de los reportes financieros — para el contador.

Genera el workbook en memoria (openpyxl) a partir de los mismos dicts que ya
devuelven los endpoints JSON, lo sube a Cloudinary (raw) y devuelve la URL:
la app la abre con el navegador/Excel del teléfono.
"""
from __future__ import annotations

import io
import uuid
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_AZUL = "0D47A1"
_HDR_FILL = PatternFill(start_color=_AZUL, end_color=_AZUL, fill_type="solid")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)


def _hoja(wb: Workbook, titulo: str, subtitulo: str,
          encabezados: list[str], filas: list[list], totales: list | None = None):
    ws = wb.active if wb.active.max_row == 1 and wb.active.max_column == 1 else wb.create_sheet()
    ws.title = titulo[:31]
    ws.append([titulo])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([subtitulo])
    ws.append([])
    ws.append(encabezados)
    hdr_row = ws.max_row
    for col in range(1, len(encabezados) + 1):
        c = ws.cell(row=hdr_row, column=col)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center")
    for fila in filas:
        ws.append(fila)
    if totales:
        ws.append(totales)
        for col in range(1, len(encabezados) + 1):
            ws.cell(row=ws.max_row, column=col).font = _BOLD
    # anchos: primera col ancha (nombres), resto numérico
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 42
    for col in range(3, len(encabezados) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    # formato numérico en columnas de montos (desde la 3ra)
    for row in ws.iter_rows(min_row=hdr_row + 1, min_col=3):
        for c in row:
            if isinstance(c.value, (int, float)):
                c.number_format = "#,##0.00"
    return ws


def _f(v) -> float:
    return float(v or 0)


def xlsx_balance_comprobacion(data: dict, empresa: str) -> bytes:
    wb = Workbook()
    _hoja(
        wb, "Balance de comprobación",
        f"{empresa} — período {data['periodo']}",
        ["Código", "Cuenta", "Débito", "Crédito", "Saldo deudor", "Saldo acreedor"],
        [[c["codigo"], c["nombre"], _f(c["total_debito"]), _f(c["total_credito"]),
          _f(c["saldo_deudor"]), _f(c["saldo_acreedor"])] for c in data["cuentas"]],
        ["", "TOTALES", "", "", _f(data["total_deudor"]), _f(data["total_acreedor"])],
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def xlsx_estado_resultados(data: dict, empresa: str) -> bytes:
    wb = Workbook()
    filas = ([["", "INGRESOS", ""]]
             + [[i["codigo"], i["nombre"], _f(i["monto"])] for i in data["ingresos"]]
             + [["", "Total ingresos", _f(data["total_ingresos"])], ["", "", ""]]
             + [["", "GASTOS", ""]]
             + [[g["codigo"], g["nombre"], _f(g["monto"])] for g in data["gastos"]]
             + [["", "Total gastos", _f(data["total_gastos"])]])
    _hoja(
        wb, "Estado de resultados",
        f"{empresa} — del {data['desde']} al {data['hasta']}",
        ["Código", "Concepto", "Monto"],
        filas,
        ["", "RESULTADO", _f(data["resultado"])],
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def xlsx_balance_general(data: dict, empresa: str) -> bytes:
    wb = Workbook()
    filas = ([["", "ACTIVO", ""]]
             + [[a["codigo"], a["nombre"], _f(a["saldo"])] for a in data["activo"]]
             + [["", "Total activo", _f(data["total_activo"])], ["", "", ""]]
             + [["", "PASIVO", ""]]
             + [[p["codigo"], p["nombre"], _f(p["saldo"])] for p in data["pasivo"]]
             + [["", "Total pasivo", _f(data["total_pasivo"])], ["", "", ""]]
             + [["", "PATRIMONIO", ""]]
             + [[p["codigo"], p["nombre"], _f(p["saldo"])] for p in data["patrimonio"]])
    _hoja(
        wb, "Balance general",
        f"{empresa} — al {data['fecha']}",
        ["Código", "Concepto", "Saldo"],
        filas,
        ["", "PASIVO + PATRIMONIO", _f(data.get("total_pasivo_patrimonio",
                                                data.get("pasivo_patrimonio", 0)))],
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def subir_xlsx(data: bytes, nombre: str) -> str:
    """Sube el XLSX a Cloudinary (raw) y devuelve la URL pública."""
    from app.services.cloudinary_service import subir_bytes_raw_cloudinary

    public_id = f"reportes/{nombre}-{date.today().isoformat()}-{uuid.uuid4().hex[:6]}"
    return subir_bytes_raw_cloudinary(data, public_id, ext="xlsx")

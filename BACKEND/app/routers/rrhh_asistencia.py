"""
HU-30 Fase 3 — Asistencia: Motor de Cálculo + Advertencias + Detalle Diario + Reportes

Endpoints:
  GET  /rrhh/asistencia/resumen                    resumen semanal por empleado (paginado)
  GET  /rrhh/asistencia/diario                     vista diaria: todos los empleados para una fecha
  GET  /rrhh/asistencia/{empleado_id}/detalle       detalle diario de registros de un empleado
  POST /rrhh/asistencia/{empleado_id}/advertencia   emite memorándum de advertencia
  GET  /rrhh/asistencia/reportes/global             reporte global todos empleados (xlsx/pdf)
  GET  /rrhh/asistencia/reportes/semanal            exporta reporte semanal (xlsx/pdf)
  GET  /rrhh/asistencia/reportes/mensual            exporta reporte mensual
  GET  /rrhh/asistencia/reportes/tardanzas          exporta incidencias de tardanzas/faltas
  GET  /rrhh/asistencia/reportes/individual         exporta historial de un empleado
  GET  /rrhh/asistencia/reportes/horas-extra        exporta empleados con >48h semanales
"""
from __future__ import annotations

import io
import re as _re
import uuid as _uuid
from datetime import date, datetime, time, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.db.database import get_db
from app.core.security import verificar_token
from app.core.permisos import exigir_no_tecnico
from app.models.registro_asistencia import RegistroAsistencia
from app.models.geolocalizacion_asistencia import GeolocalizacionAsistencia
from app.models.solicitud_laboral import SolicitudLaboral
from app.models.documento_laboral import DocumentoLaboral
from app.models.feriado_empresa import FeriadoEmpresa
from app.models.empleado import Empleado
from app.models.usuario import Usuario
from app.models.area import Area

router = APIRouter(tags=["RRHH · Asistencia"])

META_HORAS_DIA   = 8.0
META_HORAS_SEM   = 48.0
MAX_ADVERTENCIAS = 3
TIPO_MEMO        = "memorandum"
NOMBRE_MEMO      = "Memorándum de Advertencia por Faltas/Tardanzas"

# ── UUID/Área helpers ─────────────────────────────────────────────────────────
_UUID_RE = _re.compile(
    r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$',
    _re.IGNORECASE
)

def _area_cache(db: Session, empresa_id: str) -> dict:
    """Carga todas las áreas de la empresa: {id → nombre}."""
    return {a.id: a.nombre for a in db.query(Area).filter(Area.empresa_id == empresa_id).all()}


def _feriados_set(db: Session, empresa_id: str, inicio: date, fin: date) -> set[date]:
    """Devuelve el conjunto de fechas feriadas registradas para la empresa en el rango."""
    rows = (
        db.query(FeriadoEmpresa.fecha)
        .filter(
            FeriadoEmpresa.empresa_id == empresa_id,
            FeriadoEmpresa.fecha >= inicio,
            FeriadoEmpresa.fecha <= fin,
        )
        .all()
    )
    return {r.fecha for r in rows}


def _estado_dia_turno(
    registros_dia: list,
    horas: float,
    hora_entrada_turno: "time",
    tolerancia_min: int,
    req_horas: float,
) -> str:
    """Versión shift-aware de _estado_dia: evalúa contra el turno real del empleado."""
    entrada = next((r for r in registros_dia if r.tipo == "entrada"), None)
    salida  = next((r for r in registros_dia if r.tipo == "salida"),  None)

    if not entrada:
        return "Falta"

    limite = entrada.fecha_hora.replace(
        hour=hora_entrada_turno.hour,
        minute=hora_entrada_turno.minute + tolerancia_min,
        second=0, microsecond=0,
    )
    if entrada.fecha_hora > limite:
        return "Tardanza"

    if not salida or horas < req_horas - 0.5:
        return "Incompleto"

    return "Al día"

def _resolve_area(val, cache: dict) -> str:
    """Resuelve emp.area: si es UUID lo convierte al nombre real, si es texto lo devuelve tal cual."""
    if not val:
        return ""
    s = str(val).strip()
    return cache.get(s, "") if _UUID_RE.match(s) else s

# Hora tolerada de entrada (tardanza = después de las 8:00 + 10 min)
_HORA_ENTRADA = 8   # 08:00
_MIN_TOLERANCIA = 10


# ── Helpers de período ────────────────────────────────────────────────────────

def _lunes_semana(ref: date) -> date:
    return ref - timedelta(days=ref.weekday())

def _sabado_semana(ref: date) -> date:
    return _lunes_semana(ref) + timedelta(days=5)

def _dias_laborables(inicio: date, fin: date) -> list[date]:
    dias, cur = [], inicio
    while cur <= fin:
        if cur.weekday() < 6:
            dias.append(cur)
        cur += timedelta(days=1)
    return dias

_DIAS_ES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]


# ── Helpers de cálculo ────────────────────────────────────────────────────────

def _horas_dia(registros_dia: list) -> float:
    entrada  = next((r.fecha_hora for r in registros_dia if r.tipo == "entrada"),          None)
    salida   = next((r.fecha_hora for r in registros_dia if r.tipo == "salida"),           None)
    ini_alm  = next((r.fecha_hora for r in registros_dia if r.tipo == "entrada_almuerzo"), None)
    fin_alm  = next((r.fecha_hora for r in registros_dia if r.tipo == "salida_almuerzo"),  None)

    if not entrada or not salida:
        return 0.0

    horas = (salida - entrada).total_seconds() / 3600.0
    if ini_alm and fin_alm and fin_alm > ini_alm:
        horas -= (fin_alm - ini_alm).total_seconds() / 3600.0

    return max(0.0, horas)


def _estado_dia(registros_dia: list, horas: float, es_dia_laborable: bool) -> str:
    if not es_dia_laborable:
        return "—"
    entrada = next((r.fecha_hora for r in registros_dia if r.tipo == "entrada"), None)
    salida  = next((r.fecha_hora for r in registros_dia if r.tipo == "salida"),  None)

    if not entrada:
        return "Falta"

    # Tardanza: entrada después de HH:MM + tolerancia
    hora_limite = entrada.replace(
        hour=_HORA_ENTRADA, minute=_MIN_TOLERANCIA, second=0, microsecond=0)
    if entrada > hora_limite:
        return "Tardanza"

    if not salida:
        return "Incompleto"

    if horas >= META_HORAS_DIA - 0.5:  # margen de 30 min
        return "Al día"

    return "Incompleto"


def _contar_advertencias(db: Session, empleado_id: str, empresa_id: str) -> int:
    return (
        db.query(DocumentoLaboral)
        .filter(
            DocumentoLaboral.empleado_id == empleado_id,
            DocumentoLaboral.empresa_id  == empresa_id,
            DocumentoLaboral.tipo        == TIPO_MEMO,
        )
        .count()
    )


# ── 1. GET /rrhh/asistencia/resumen ──────────────────────────────────────────

@router.get("/rrhh/asistencia/resumen")
def resumen_asistencia(
    fecha_inicio: Optional[date] = Query(None),
    fecha_fin:    Optional[date] = Query(None),
    page:         int            = Query(1,  ge=1),
    limit:        int            = Query(10, ge=1, le=100),
    db:           Session        = Depends(get_db),
    payload:      dict           = Depends(verificar_token),
):
    exigir_no_tecnico(payload)
    empresa_id = payload["empresa_id"]

    hoy    = date.today()
    inicio = fecha_inicio or _lunes_semana(hoy)
    fin    = fecha_fin    or _sabado_semana(hoy)

    if fin < inicio:
        raise HTTPException(400, "fecha_fin debe ser >= fecha_inicio")

    dias_lab     = _dias_laborables(inicio, fin)
    dias_lab_set = set(dias_lab)
    meta_horas   = len(dias_lab) * META_HORAS_DIA

    empleados_q = (
        db.query(Empleado, Usuario)
        .join(Usuario, Usuario.id == Empleado.usuario_id)
        .filter(Empleado.empresa_id == empresa_id, Empleado.activo == True)
    )
    total_empleados = empleados_q.count()

    empleados = (
        empleados_q
        .order_by(Usuario.nombre, Usuario.apellido)
        .all()
    )

    inicio_dt = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)
    fin_dt    = datetime(fin.year,    fin.month,    fin.day,    23, 59, 59)

    todos_registros = (
        db.query(RegistroAsistencia)
        .filter(
            RegistroAsistencia.empresa_id == empresa_id,
            RegistroAsistencia.fecha_hora >= inicio_dt,
            RegistroAsistencia.fecha_hora <= fin_dt,
        )
        .all()
    )

    todas_solicitudes = (
        db.query(SolicitudLaboral)
        .filter(
            SolicitudLaboral.empresa_id  == empresa_id,
            SolicitudLaboral.estado      == "aprobada",
            SolicitudLaboral.fecha_inicio <= fin,
            SolicitudLaboral.fecha_fin   >= inicio,
        )
        .all()
    )

    regs_por_emp:  dict[str, list] = {}
    for reg in todos_registros:
        regs_por_emp.setdefault(reg.empleado_id, []).append(reg)

    solis_por_emp: dict[str, list] = {}
    for sol in todas_solicitudes:
        solis_por_emp.setdefault(sol.empleado_id, []).append(sol)

    area_nombres = _area_cache(db, empresa_id)
    resultado = []

    for emp, usr in empleados:
        regs  = regs_por_emp.get(emp.id,  [])
        solis = solis_por_emp.get(emp.id, [])

        dias_justificados: set[date] = set()
        for sol in solis:
            cur = max(sol.fecha_inicio, inicio)
            end = min(sol.fecha_fin,    fin)
            while cur <= end:
                if cur in dias_lab_set:
                    dias_justificados.add(cur)
                cur += timedelta(days=1)

        horas_reales = 0.0
        dias_laborados = 0
        for dia in dias_lab:
            if dia not in dias_justificados:
                regs_dia = [r for r in regs if r.fecha_hora.date() == dia]
                h = _horas_dia(regs_dia)
                horas_reales += h
                if h > 0:
                    dias_laborados += 1

        horas_justificadas = len(dias_justificados) * META_HORAS_DIA
        horas_total        = horas_reales + horas_justificadas
        horas_faltantes    = max(0.0, meta_horas - horas_total)
        porcentaje         = round((horas_total / meta_horas * 100) if meta_horas > 0 else 0.0, 1)

        advertencias = _contar_advertencias(db, emp.id, empresa_id)

        nombre    = f"{usr.nombre} {usr.apellido}".strip()
        iniciales = (
            (usr.nombre[0]   if usr.nombre   else "") +
            (usr.apellido[0] if usr.apellido else "")
        ).upper() or "?"

        resultado.append({
            "id":                 emp.id,
            "nombreCompleto":     nombre,
            "cargo":              emp.cargo,
            "area":               _resolve_area(emp.area, area_nombres),
            "iniciales":          iniciales,
            "fotoUrl":            usr.foto_url or "",
            "tipo_contrato":      emp.tipo,
            "codigo":             emp.codigo,
            "tipo_documento":     emp.tipo_documento or "DNI",
            "numero_documento":   emp.numero_documento,
            "cuspp":              emp.cuspp,
            "fecha_ingreso":      emp.fecha_ingreso.isoformat() if emp.fecha_ingreso else None,
            "horas_reales":       round(horas_reales,       2),
            "horas_justificadas": round(horas_justificadas, 2),
            "horas_total":        round(horas_total,        2),
            "horas_faltantes":    round(horas_faltantes,    2),
            "dias_laborados":     dias_laborados,
            "meta_horas":         meta_horas,
            "porcentaje":         porcentaje,
            "advertencias":       advertencias,
        })

    resultado.sort(key=lambda e: e["horas_faltantes"], reverse=True)

    # Paginación en memoria sobre el resultado ya ordenado
    total = len(resultado)
    offset = (page - 1) * limit
    pagina = resultado[offset: offset + limit]

    return {
        "periodo": {
            "fecha_inicio": inicio.isoformat(),
            "fecha_fin":    fin.isoformat(),
            "meta_horas":   meta_horas,
        },
        "empleados":     pagina,
        "total":         total,
        "page":          page,
        "limit":         limit,
        "total_paginas": max(1, -(-total // limit)),
    }


# ── 1b. GET /rrhh/asistencia/diario ─────────────────────────────────────────
# NOTA: este endpoint DEBE definirse ANTES de /{empleado_id}/detalle para que
#       FastAPI no interprete "diario" como un empleado_id.

@router.get("/rrhh/asistencia/diario")
def asistencia_diaria(
    fecha:   date = Query(...),
    page:    int  = Query(1,  ge=1),
    limit:   int  = Query(10, ge=1, le=100),
    db:      Session = Depends(get_db),
    payload: dict    = Depends(verificar_token),
):
    """Vista diaria: todos los empleados activos para la fecha indicada."""
    exigir_no_tecnico(payload)
    empresa_id = payload["empresa_id"]

    fecha_dt_ini = datetime(fecha.year, fecha.month, fecha.day, 0,  0,  0)
    fecha_dt_fin = datetime(fecha.year, fecha.month, fecha.day, 23, 59, 59)

    empleados_q = (
        db.query(Empleado, Usuario)
        .join(Usuario, Usuario.id == Empleado.usuario_id)
        .filter(Empleado.empresa_id == empresa_id, Empleado.activo == True)
        .order_by(Usuario.nombre, Usuario.apellido)
        .all()
    )
    total = len(empleados_q)
    area_nombres = _area_cache(db, empresa_id)

    registros = (
        db.query(RegistroAsistencia)
        .filter(
            RegistroAsistencia.empresa_id == empresa_id,
            RegistroAsistencia.fecha_hora >= fecha_dt_ini,
            RegistroAsistencia.fecha_hora <= fecha_dt_fin,
        )
        .all()
    )

    reg_ids = [r.id for r in registros]
    geos: dict[str, GeolocalizacionAsistencia] = {}
    if reg_ids:
        for g in db.query(GeolocalizacionAsistencia).filter(
            GeolocalizacionAsistencia.registro_id.in_(reg_ids)
        ).all():
            geos[g.registro_id] = g

    regs_por_emp: dict[str, list] = {}
    for r in registros:
        regs_por_emp.setdefault(r.empleado_id, []).append(r)

    es_lab = fecha.weekday() < 6  # Lun-Sáb

    def _geo(reg):
        if not reg:
            return None
        g = geos.get(reg.id)
        if not g:
            return None
        return {"lat": float(g.latitud), "lng": float(g.longitud),
                "precision": float(g.precision_m) if g.precision_m else None}

    def _lat(reg):
        g = geos.get(reg.id) if reg else None
        return float(g.latitud) if g else None

    def _lng(reg):
        g = geos.get(reg.id) if reg else None
        return float(g.longitud) if g else None

    filas = []
    for emp, usr in empleados_q:
        regs = regs_por_emp.get(emp.id, [])

        entrada  = next((r for r in regs if r.tipo == "entrada"),          None)
        salida   = next((r for r in regs if r.tipo == "salida"),           None)
        ini_alm  = next((r for r in regs if r.tipo == "entrada_almuerzo"), None)
        fin_alm  = next((r for r in regs if r.tipo == "salida_almuerzo"),  None)

        horas  = _horas_dia(regs)
        estado = _estado_dia(regs, horas, es_lab)

        nombre    = f"{usr.nombre} {usr.apellido}".strip()
        iniciales = (
            (usr.nombre[0]   if usr.nombre   else "") +
            (usr.apellido[0] if usr.apellido else "")
        ).upper() or "?"

        filas.append({
            "empleado_id":     emp.id,
            "nombreCompleto":  nombre,
            "cargo":           emp.cargo or "",
            "area":            _resolve_area(emp.area, area_nombres),
            "iniciales":       iniciales,
            "fotoUrl":         usr.foto_url or "",
            "hora_ingreso":    entrada.fecha_hora.strftime("%H:%M") if entrada  else None,
            "hora_salida":     salida.fecha_hora.strftime("%H:%M")  if salida   else None,
            "almuerzo_inicio": ini_alm.fecha_hora.strftime("%H:%M") if ini_alm  else None,
            "almuerzo_fin":    fin_alm.fecha_hora.strftime("%H:%M") if fin_alm  else None,
            "horas_trabajadas": round(horas, 2),
            "lat_ingreso":     _lat(entrada),
            "lng_ingreso":     _lng(entrada),
            "lat_salida":      _lat(salida),
            "lng_salida":      _lng(salida),
            "geo_ingreso":     _geo(entrada),
            "geo_salida":      _geo(salida),
            "estado":          estado,
        })

    offset = (page - 1) * limit
    pagina = filas[offset: offset + limit]

    return {
        "fecha":         fecha.isoformat(),
        "empleados":     pagina,
        "total":         total,
        "page":          page,
        "limit":         limit,
        "total_paginas": max(1, -(-total // limit)),
    }


# ── 2. GET /rrhh/asistencia/{empleado_id}/detalle ────────────────────────────

@router.get("/rrhh/asistencia/{empleado_id}/detalle")
def detalle_diario(
    empleado_id: str,
    fecha_inicio: Optional[date] = Query(None),
    fecha_fin:    Optional[date] = Query(None),
    page:         int            = Query(1,  ge=1),
    limit:        int            = Query(10, ge=1, le=100),
    db:           Session        = Depends(get_db),
    payload:      dict           = Depends(verificar_token),
):
    exigir_no_tecnico(payload)
    empresa_id = payload["empresa_id"]

    emp = db.query(Empleado).filter(
        Empleado.id == empleado_id, Empleado.empresa_id == empresa_id
    ).first()
    if not emp:
        raise HTTPException(404, "Empleado no encontrado")

    hoy    = date.today()
    inicio = fecha_inicio or _lunes_semana(hoy)
    fin    = fecha_fin    or _sabado_semana(hoy)

    if fin < inicio:
        raise HTTPException(400, "fecha_fin debe ser >= fecha_inicio")

    dias_lab = _dias_laborables(inicio, fin)

    inicio_dt = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)
    fin_dt    = datetime(fin.year,    fin.month,    fin.day,    23, 59, 59)

    registros = (
        db.query(RegistroAsistencia)
        .filter(
            RegistroAsistencia.empresa_id  == empresa_id,
            RegistroAsistencia.empleado_id == empleado_id,
            RegistroAsistencia.fecha_hora  >= inicio_dt,
            RegistroAsistencia.fecha_hora  <= fin_dt,
        )
        .all()
    )

    # Cargar geolocalizaciones
    reg_ids = [r.id for r in registros]
    geos: dict[str, GeolocalizacionAsistencia] = {}
    if reg_ids:
        for g in db.query(GeolocalizacionAsistencia).filter(
            GeolocalizacionAsistencia.registro_id.in_(reg_ids)
        ).all():
            geos[g.registro_id] = g

    # Agregar por día
    regs_por_dia: dict[date, list] = {}
    for r in registros:
        d = r.fecha_hora.date()
        regs_por_dia.setdefault(d, []).append(r)

    filas = []
    for dia in sorted(set(dias_lab) | set(regs_por_dia.keys()), reverse=True):
        regs_dia = regs_por_dia.get(dia, [])
        es_lab   = dia in set(dias_lab)

        entrada   = next((r for r in regs_dia if r.tipo == "entrada"),          None)
        salida    = next((r for r in regs_dia if r.tipo == "salida"),           None)
        ini_alm   = next((r for r in regs_dia if r.tipo == "entrada_almuerzo"), None)
        fin_alm   = next((r for r in regs_dia if r.tipo == "salida_almuerzo"),  None)

        horas = _horas_dia(regs_dia)
        estado = _estado_dia(regs_dia, horas, es_lab)

        def _geo_dict(reg):
            if not reg:
                return None
            g = geos.get(reg.id)
            if not g:
                return None
            return {
                "lat":      float(g.latitud),
                "lng":      float(g.longitud),
                "precision": float(g.precision_m) if g.precision_m else None,
            }

        def _fmt_time(dt):
            return dt.strftime("%H:%M") if dt else None

        dur_alm = None
        if ini_alm and fin_alm and fin_alm > ini_alm:
            mins = int((fin_alm - ini_alm).total_seconds() / 60)
            dur_alm = f"{mins} min"

        filas.append({
            "fecha":          dia.isoformat(),
            "dia_nombre":     _DIAS_ES[dia.weekday()],
            "hora_ingreso":   _fmt_time(entrada.fecha_hora  if entrada  else None),
            "hora_salida":    _fmt_time(salida.fecha_hora   if salida   else None),
            "almuerzo_inicio":_fmt_time(ini_alm.fecha_hora  if ini_alm  else None),
            "almuerzo_fin":   _fmt_time(fin_alm.fecha_hora  if fin_alm  else None),
            "almuerzo_dur":   dur_alm,
            "horas_trabajadas": round(horas, 2),
            "estado":          estado,
            "geo_ingreso":     _geo_dict(entrada),
            "geo_salida":      _geo_dict(salida),
        })

    total = len(filas)
    offset = (page - 1) * limit
    pagina = filas[offset: offset + limit]

    return {
        "empleado_id": empleado_id,
        "periodo": {"fecha_inicio": inicio.isoformat(), "fecha_fin": fin.isoformat()},
        "dias": pagina,
        "total": total,
        "page": page,
        "limit": limit,
        "total_paginas": max(1, -(-total // limit)),
    }


# ── 3. POST /rrhh/asistencia/{empleado_id}/advertencia ───────────────────────

@router.post("/rrhh/asistencia/{empleado_id}/advertencia", status_code=201)
def emitir_advertencia(
    empleado_id: str,
    db:      Session = Depends(get_db),
    payload: dict    = Depends(verificar_token),
):
    exigir_no_tecnico(payload, "Solo personal de RRHH puede emitir advertencias")
    empresa_id = payload["empresa_id"]

    emp = (
        db.query(Empleado)
        .filter(
            Empleado.id         == empleado_id,
            Empleado.empresa_id == empresa_id,
            Empleado.activo     == True,
        )
        .first()
    )
    if not emp:
        raise HTTPException(404, "Empleado no encontrado")

    n = _contar_advertencias(db, emp.id, empresa_id)

    if n >= MAX_ADVERTENCIAS:
        raise HTTPException(
            400,
            "Límite legal de advertencias alcanzado (Falta grave). "
            "El empleado ya tiene 3 memorándums emitidos.",
        )

    memo = DocumentoLaboral(
        id                   = str(_uuid.uuid4()),
        empleado_id          = emp.id,
        empresa_id           = empresa_id,
        tipo                 = TIPO_MEMO,
        nombre               = NOMBRE_MEMO,
        url_archivo          = None,
        public_id_cloudinary = None,
        fecha_emision        = date.today(),
        requiere_firma       = False,
    )
    db.add(memo)
    db.commit()
    db.refresh(memo)

    nueva_cuenta = n + 1
    return {
        "id":                    memo.id,
        "mensaje":               f"Advertencia N°{nueva_cuenta} registrada correctamente.",
        "numero_advertencia":    nueva_cuenta,
        "advertencias_restantes": MAX_ADVERTENCIAS - nueva_cuenta,
    }


# ══════════════════════════════════════════════════════════════════════════════
# REPORTES DE ASISTENCIA
# ══════════════════════════════════════════════════════════════════════════════

def _build_resumen_full(db: Session, empresa_id: str, inicio: date, fin: date):
    """
    Construye todos los datos del resumen para reportes (sin paginar).
    — Respeta el turno real de cada empleado (shift-aware).
    — Excluye feriados de empresa: se tratan como días no laborables.
    """
    from app.models.turno import Turno, TurnoEmpleado

    dias_lab     = _dias_laborables(inicio, fin)
    dias_lab_set = set(dias_lab)

    # Feriados registrados → se excluyen del cómputo de faltas
    feriados = _feriados_set(db, empresa_id, inicio, fin)
    # Días laborables efectivos (sin feriados)
    dias_lab_ef  = [d for d in dias_lab if d not in feriados]
    meta_horas   = len(dias_lab_ef) * META_HORAS_DIA

    empleados = (
        db.query(Empleado, Usuario)
        .join(Usuario, Usuario.id == Empleado.usuario_id)
        .filter(Empleado.empresa_id == empresa_id, Empleado.activo == True)
        .order_by(Usuario.nombre, Usuario.apellido)
        .all()
    )

    inicio_dt = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)
    fin_dt    = datetime(fin.year,    fin.month,    fin.day,    23, 59, 59)

    todos_registros = (
        db.query(RegistroAsistencia)
        .filter(
            RegistroAsistencia.empresa_id == empresa_id,
            RegistroAsistencia.fecha_hora >= inicio_dt,
            RegistroAsistencia.fecha_hora <= fin_dt,
        )
        .all()
    )

    todas_solicitudes = (
        db.query(SolicitudLaboral)
        .filter(
            SolicitudLaboral.empresa_id  == empresa_id,
            SolicitudLaboral.estado      == "aprobada",
            SolicitudLaboral.fecha_inicio <= fin,
            SolicitudLaboral.fecha_fin   >= inicio,
        )
        .all()
    )

    # Asignaciones de turno vigentes en el rango (una sola consulta)
    asigns = (
        db.query(TurnoEmpleado, Turno)
        .join(Turno, Turno.id == TurnoEmpleado.turno_id)
        .filter(
            Turno.empresa_id == empresa_id,
            TurnoEmpleado.activo == True, Turno.activo == True,
            TurnoEmpleado.fecha_desde <= fin,
            or_(TurnoEmpleado.fecha_hasta.is_(None), TurnoEmpleado.fecha_hasta >= inicio),
        ).all()
    )
    asigns_por_emp: dict[str, list] = {}
    for te, turno in asigns:
        asigns_por_emp.setdefault(str(te.empleado_id), []).append((te, turno))

    def _turno_emp_dia(emp_id: str, dia: date):
        """Devuelve (hora_entrada, tolerancia, horas_requeridas) para el día."""
        for te, turno in asigns_por_emp.get(emp_id, []):
            if te.fecha_desde <= dia and (te.fecha_hasta is None or te.fecha_hasta >= dia):
                req_min = max(
                    _min_entre_horas(turno.hora_entrada, turno.hora_salida)
                    - (turno.duracion_almuerzo_minutos or 0), 0
                )
                return (
                    turno.hora_entrada,
                    turno.tolerancia_minutos if turno.tolerancia_minutos is not None
                    else _CRONO_TOLERANCIA,
                    req_min / 60.0,
                )
        return (_CRONO_ENTRADA, _CRONO_TOLERANCIA, META_HORAS_DIA)

    regs_por_emp:  dict[str, list] = {}
    for r in todos_registros:
        regs_por_emp.setdefault(r.empleado_id, []).append(r)

    solis_por_emp: dict[str, list] = {}
    for s in todas_solicitudes:
        solis_por_emp.setdefault(s.empleado_id, []).append(s)

    area_nombres = _area_cache(db, empresa_id)
    filas = []
    for emp, usr in empleados:
        regs  = regs_por_emp.get(emp.id,  [])
        solis = solis_por_emp.get(emp.id, [])

        dias_justificados: set[date] = set()
        for sol in solis:
            cur = max(sol.fecha_inicio, inicio)
            end = min(sol.fecha_fin,    fin)
            while cur <= end:
                if cur in dias_lab_set:
                    dias_justificados.add(cur)
                cur += timedelta(days=1)

        horas_reales = 0.0
        faltas = tardanzas = 0

        for dia in dias_lab_ef:                     # excluye feriados
            if dia in dias_justificados:
                horas_reales += META_HORAS_DIA      # justificado cuenta como cumplido
                continue
            regs_dia = [r for r in regs if r.fecha_hora.date() == dia]
            h = _horas_dia(regs_dia)
            horas_reales += h
            t_ent, t_tol, req_h = _turno_emp_dia(emp.id, dia)
            estado = _estado_dia_turno(regs_dia, h, t_ent, t_tol, req_h)
            if estado == "Falta":
                faltas += 1
            elif estado == "Tardanza":
                tardanzas += 1

        # Horas justificadas = días con solicitud aprobada × horas del turno ese día
        horas_justificadas = sum(
            _turno_emp_dia(emp.id, d)[2]
            for d in dias_justificados
            if d in dias_lab_ef
        )
        horas_total     = horas_reales
        horas_faltantes = max(0.0, meta_horas - horas_total)
        horas_extra     = max(0.0, horas_total - meta_horas)
        porcentaje      = round((horas_total / meta_horas * 100) if meta_horas > 0 else 0.0, 1)

        filas.append({
            "nombre":             f"{usr.nombre} {usr.apellido}".strip(),
            "cargo":              emp.cargo or "",
            "area":               _resolve_area(emp.area, area_nombres),
            "horas_reales":       round(horas_reales,       2),
            "horas_justificadas": round(horas_justificadas, 2),
            "horas_total":        round(horas_total,        2),
            "horas_faltantes":    round(horas_faltantes,    2),
            "horas_extra":        round(horas_extra,        2),
            "meta_horas":         meta_horas,
            "porcentaje":         porcentaje,
            "faltas":             faltas,
            "tardanzas":          tardanzas,
        })

    return filas, meta_horas


# ── Cronograma (matriz día × empleado) para el PDF mensual/semanal ───────────
# Código de celda → color de fondo / etiqueta de leyenda.
_CRONO_COLOR = {
    "P": "#DCFCE7",  # Presente / al día     → verde
    "T": "#FEF3C7",  # Tardanza              → amarillo
    "F": "#FEE2E2",  # Falta                 → rojo
    "I": "#FFEDD5",  # Incompleto            → naranja
    "J": "#DBEAFE",  # Justificado           → azul
    "D": "#F3F4F6",  # Descanso / domingo    → gris
    "H": "#EDE9FE",  # Feriado de empresa    → violeta
}
_CRONO_LABEL = {
    "P": "Presente", "T": "Tardanza", "F": "Falta",
    "I": "Incompleto", "J": "Justificado", "D": "Descanso", "H": "Feriado",
}


# Turno por defecto (sin excepción asignada). Tolerancia 10 min: la entrada se
# considera puntual hasta 08:10; recién a partir de 08:11 cuenta como tardanza.
_CRONO_ENTRADA   = time(8, 0)
_CRONO_SALIDA    = time(17, 0)
_CRONO_ALMUERZO  = 60
_CRONO_TOLERANCIA = 10


def _min_entre_horas(t1: time, t2: time) -> int:
    """Minutos (t2 - t1) en el mismo día; nunca negativo."""
    d = (t2.hour * 60 + t2.minute) - (t1.hour * 60 + t1.minute)
    return d if d > 0 else 0


def _build_cronograma(db: Session, empresa_id: str, inicio: date, fin: date):
    """Devuelve (dias, empleados, totales) para el cronograma de asistencia.

    SHIFT-AWARE: la puntualidad de cada empleado se mide contra SU turno
    (excepción de horario vigente o el turno por defecto 08:00–17:00, tol.
    10 min). Cada empleado trae KPIs, minutos de tardanza por día (para el
    gráfico de su hoja-perfil) y las justificaciones aprobadas del período.

    dias      : TODAS las fechas del rango (columnas; domingos = descanso).
    empleados : ver estructura del dict construido abajo.
    totales   : agregados del período (incluye KPIs de puntualidad).
    """
    from ..models.turno import Turno, TurnoEmpleado

    dias_all: list[date] = []
    cur = inicio
    while cur <= fin:
        dias_all.append(cur)
        cur += timedelta(days=1)
    dias_lab_set = set(_dias_laborables(inicio, fin))

    empleados = (
        db.query(Empleado, Usuario)
        .join(Usuario, Usuario.id == Empleado.usuario_id)
        .filter(Empleado.empresa_id == empresa_id, Empleado.activo == True)
        .order_by(Usuario.nombre, Usuario.apellido)
        .all()
    )

    inicio_dt = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)
    fin_dt    = datetime(fin.year, fin.month, fin.day, 23, 59, 59)

    registros = (
        db.query(RegistroAsistencia)
        .filter(
            RegistroAsistencia.empresa_id == empresa_id,
            RegistroAsistencia.fecha_hora >= inicio_dt,
            RegistroAsistencia.fecha_hora <= fin_dt,
        ).all()
    )
    solicitudes = (
        db.query(SolicitudLaboral)
        .filter(
            SolicitudLaboral.empresa_id   == empresa_id,
            SolicitudLaboral.estado       == "aprobada",
            SolicitudLaboral.fecha_inicio <= fin,
            SolicitudLaboral.fecha_fin    >= inicio,
        ).all()
    )
    # Asignaciones de turno-excepción vigentes en el rango (una sola consulta).
    asigns = (
        db.query(TurnoEmpleado, Turno)
        .join(Turno, Turno.id == TurnoEmpleado.turno_id)
        .filter(
            Turno.empresa_id == empresa_id,
            TurnoEmpleado.activo == True, Turno.activo == True,
            TurnoEmpleado.fecha_desde <= fin,
            or_(TurnoEmpleado.fecha_hasta.is_(None), TurnoEmpleado.fecha_hasta >= inicio),
        ).all()
    )
    asigns_por_emp: dict[str, list] = {}
    for te, turno in asigns:
        asigns_por_emp.setdefault(str(te.empleado_id), []).append((te, turno))

    def _turno_dia(emp_id, dia: date):
        """(hora_entrada, hora_salida, almuerzo_min, tolerancia, nombre) del día.
        Normaliza emp_id a str: las claves del dict son str(empleado_id) y los
        ids de SQLAlchemy pueden venir como UUID nativo (no matchearían)."""
        for te, turno in asigns_por_emp.get(str(emp_id), []):
            if te.fecha_desde <= dia and (te.fecha_hasta is None or te.fecha_hasta >= dia):
                return (turno.hora_entrada, turno.hora_salida,
                        turno.duracion_almuerzo_minutos or 0,
                        turno.tolerancia_minutos if turno.tolerancia_minutos is not None
                        else _CRONO_TOLERANCIA, turno.nombre)
        return (_CRONO_ENTRADA, _CRONO_SALIDA, _CRONO_ALMUERZO,
                _CRONO_TOLERANCIA, "Horario normal")

    # Feriados registrados para la empresa en el rango
    feriados_crono = _feriados_set(db, empresa_id, inicio, fin)

    regs_por_emp: dict[str, list] = {}
    for r in registros:
        regs_por_emp.setdefault(r.empleado_id, []).append(r)
    solis_por_emp: dict[str, list] = {}
    for s in solicitudes:
        solis_por_emp.setdefault(s.empleado_id, []).append(s)

    area_nombres = _area_cache(db, empresa_id)
    empleados_data = []
    tot_faltas = tot_tardanzas = tot_dias_trab = tot_min_tarde = 0
    suma_pct = 0.0
    for emp, usr in empleados:
        regs  = regs_por_emp.get(emp.id, [])
        solis = solis_por_emp.get(emp.id, [])

        dias_justificados: set[date] = set()
        for sol in solis:
            c = max(sol.fecha_inicio, inicio)
            e = min(sol.fecha_fin, fin)
            while c <= e:
                dias_justificados.add(c)
                c += timedelta(days=1)

        codigos: list[str] = []
        mins_tarde_dia: list[int] = []   # paralelo a dias_all (0 si no aplica)
        dias_trab = faltas = tardanzas = incompletos = 0
        min_tarde_total = 0
        horas_reales = 0.0
        min_esperados = 0
        turno_label = "Horario normal · 08:00–17:00 (tol. 10 min)"
        for dia in dias_all:
            t_ent, t_sal, t_alm, t_tol, t_nom = _turno_dia(emp.id, dia)
            if dia not in dias_lab_set:
                codigos.append("D"); mins_tarde_dia.append(0)
                continue
            # Feriado de empresa → "H" (no cuenta como falta ni como día esperado)
            if dia in feriados_crono:
                codigos.append("H"); mins_tarde_dia.append(0)
                continue
            req_min = max(_min_entre_horas(t_ent, t_sal) - t_alm, 0)
            min_esperados += req_min
            if t_nom != "Horario normal":
                turno_label = (f"{t_nom} · {t_ent.strftime('%H:%M')}–"
                               f"{t_sal.strftime('%H:%M')} (tol. {t_tol} min)")
            if dia in dias_justificados:
                codigos.append("J"); mins_tarde_dia.append(0)
                horas_reales += req_min / 60.0   # justificado cuenta como cumplido
                continue
            regs_dia = [r for r in regs if r.fecha_hora.date() == dia]
            entrada = next((r for r in regs_dia if r.tipo == "entrada"), None)
            salida  = next((r for r in regs_dia if r.tipo == "salida"), None)
            h = _horas_dia(regs_dia)
            horas_reales += h
            if not entrada:
                codigos.append("F"); mins_tarde_dia.append(0); faltas += 1
                continue
            mt = _min_entre_horas(t_ent, entrada.fecha_hora.time())
            if mt > t_tol:
                codigos.append("T"); mins_tarde_dia.append(mt)
                tardanzas += 1; dias_trab += 1
                min_tarde_total += mt
            elif not salida or h < (req_min / 60.0) - 0.5:
                codigos.append("I"); mins_tarde_dia.append(0)
                incompletos += 1; dias_trab += 1
            else:
                codigos.append("P"); mins_tarde_dia.append(0); dias_trab += 1

        pct = round((horas_reales / (min_esperados / 60.0) * 100)
                    if min_esperados > 0 else 0.0, 1)
        prom_tarde = round(min_tarde_total / tardanzas, 1) if tardanzas else 0.0
        pct_punt = round((dias_trab - tardanzas) / dias_trab * 100, 1) if dias_trab else 100.0
        suma_pct += pct
        tot_faltas += faltas; tot_tardanzas += tardanzas
        tot_dias_trab += dias_trab; tot_min_tarde += min_tarde_total

        justificaciones = [{
            "tipo":  (s.tipo or "Permiso").replace("_", " ").title(),
            "desc":  (s.descripcion or s.observacion or "").strip(),
            "desde": s.fecha_inicio, "hasta": s.fecha_fin,
        } for s in sorted(solis, key=lambda x: x.fecha_inicio or inicio)]

        empleados_data.append({
            "nombre":       f"{usr.nombre} {usr.apellido}".strip(),
            "cargo":        emp.cargo or "",
            "area":         _resolve_area(emp.area, area_nombres),
            "turno_label":  turno_label,
            "codigos":      codigos,
            "mins_tarde":   mins_tarde_dia,
            "dias_trab":    dias_trab,
            "faltas":       faltas,
            "tardanzas":    tardanzas,
            "incompletos":  incompletos,
            "min_tarde_total": min_tarde_total,
            "prom_tarde":   prom_tarde,
            "horas_reales": round(horas_reales, 1),
            "horas_esper":  round(min_esperados / 60.0, 1),
            "porcentaje":   pct,
            "pct_punt":     pct_punt,
            "justificaciones": justificaciones,
        })

    totales = {
        "n_emp":      len(empleados_data),
        "pct_global": round(suma_pct / len(empleados_data), 1) if empleados_data else 0.0,
        "faltas":     tot_faltas,
        "tardanzas":  tot_tardanzas,
        "prom_tarde_global": round(tot_min_tarde / tot_tardanzas, 1) if tot_tardanzas else 0.0,
        "punt_global": round((tot_dias_trab - tot_tardanzas) / tot_dias_trab * 100, 1)
                       if tot_dias_trab else 100.0,
    }
    return dias_all, empleados_data, totales


def _xlsx_stream(titulo: str, encabezados: list[str], filas: list[list]) -> io.BytesIO:
    """Genera un XLSX en memoria con formato básico."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(encabezados))
    cell_titulo = ws.cell(row=1, column=1, value=titulo)
    cell_titulo.font      = Font(bold=True, size=13, color="FFFFFF")
    cell_titulo.fill      = PatternFill("solid", fgColor="16A34A")
    cell_titulo.alignment = Alignment(horizontal="center")

    # Encabezados
    header_fill = PatternFill("solid", fgColor="BBF7D0")
    for col, h in enumerate(encabezados, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = Font(bold=True)
        c.fill      = header_fill
        c.alignment = Alignment(horizontal="center")

    # Datos
    for row_i, fila in enumerate(filas, start=3):
        for col_i, val in enumerate(fila, start=1):
            ws.cell(row=row_i, column=col_i, value=val)

    # Auto-ancho de columnas
    for col_i, h in enumerate(encabezados, start=1):
        max_len = max(
            len(str(h)),
            max((len(str(f[col_i - 1])) for f in filas if f[col_i - 1] is not None), default=0),
        )
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_i)
        ].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _pdf_stream(titulo: str, encabezados: list[str], filas: list[list],
                subtitulo: str = "") -> io.BytesIO:
    """Genera un PDF en memoria usando reportlab."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    story = [
        Paragraph(titulo,    styles["Title"]),
        Paragraph(subtitulo, styles["Normal"]) if subtitulo else Spacer(1, 0),
        Spacer(1, 0.4*cm),
    ]

    data = [encabezados] + [[str(v) if v is not None else "" for v in fila] for fila in filas]
    n_cols = len(encabezados)
    col_width = (doc.width) / n_cols

    t = Table(data, colWidths=[col_width] * n_cols, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#16A34A")),
        ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
        ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE",     (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0FDF4")]),
        ("GRID",         (0, 0), (-1, -1), 0.5, colors.HexColor("#D1FAE5")),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
    ]))
    story.append(t)

    doc.build(story)
    buf.seek(0)
    return buf


def _pdf_cronograma(titulo: str, subtitulo: str, dias: list[date],
                    empleados: list[dict], totales: dict,
                    detalle_por_empleado: bool = False) -> io.BytesIO:
    """PDF de asistencia organizado: resumen + KPIs + cronograma (matriz
    día×empleado con colores) + detalle por empleado. Si `detalle_por_empleado`
    (informe mensual), añade además UNA HOJA-PERFIL por trabajador con sus KPIs,
    gráfico de tardanzas y justificaciones."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, PageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.0*cm, rightMargin=1.0*cm,
                            topMargin=1.1*cm, bottomMargin=1.1*cm)
    styles = getSampleStyleSheet()
    small  = ParagraphStyle("small", parent=styles["Normal"], fontSize=8)

    story = [Paragraph(titulo, styles["Title"])]
    if subtitulo:
        story.append(Paragraph(subtitulo, styles["Normal"]))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        f"<b>Empleados:</b> {totales['n_emp']} &nbsp;&nbsp; "
        f"<b>Asistencia global:</b> {totales['pct_global']}% &nbsp;&nbsp; "
        f"<b>Puntualidad:</b> {totales.get('punt_global', 0)}% &nbsp;&nbsp; "
        f"<b>Prom. tardanza:</b> {totales.get('prom_tarde_global', 0)} min &nbsp;&nbsp; "
        f"<b>Total tardanzas:</b> {totales['tardanzas']} &nbsp;&nbsp; "
        f"<b>Total faltas:</b> {totales['faltas']}", small))
    story.append(Spacer(1, 0.35*cm))

    # ── Cronograma (matriz) ──
    story.append(Paragraph("<b>Cronograma del período</b>", styles["Heading3"]))
    n_dias = len(dias)
    header = ["Empleado"] + [str(d.day) for d in dias]
    data   = [header] + [[e["nombre"]] + e["codigos"] for e in empleados]
    name_w = 4.0*cm
    day_w  = (doc.width - name_w) / n_dias if n_dias else 0.5*cm
    t = Table(data, colWidths=[name_w] + [day_w]*n_dias, repeatRows=1)
    st = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16A34A")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME",   (0, 1), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 6),
        ("ALIGN",      (1, 0), (-1, -1), "CENTER"),
        ("ALIGN",      (0, 0), (0, -1), "LEFT"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",       (0, 0), (-1, -1), 0.3, colors.HexColor("#D1D5DB")),
        ("TOPPADDING", (0, 0), (-1, -1), 1.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        ("LEFTPADDING",  (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
    ]
    for ri, e in enumerate(empleados, start=1):
        for ci, code in enumerate(e["codigos"], start=1):
            col = _CRONO_COLOR.get(code)
            if col:
                st.append(("BACKGROUND", (ci, ri), (ci, ri), colors.HexColor(col)))
    t.setStyle(TableStyle(st))
    story.append(t)
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        "<b>Leyenda:</b> " + "&nbsp;&nbsp; ".join(
            f"{k} = {v}" for k, v in _CRONO_LABEL.items()), small))
    story.append(Spacer(1, 0.45*cm))

    # ── Detalle por empleado: resumen + incidencias del período inline ──
    # Cada trabajador ocupa UNA fila con su resumen y sus incidencias (fechas
    # compactas: "05 T · 12 F"). Reemplaza la antigua lista global de incidencias.
    story.append(Paragraph("<b>Detalle por empleado</b>", styles["Heading3"]))
    cell   = ParagraphStyle("cell",  parent=styles["Normal"], fontSize=7.5, leading=9)
    cell_b = ParagraphStyle("cellb", parent=cell, textColor=colors.white,
                            fontName="Helvetica-Bold")
    enc = [Paragraph(h, cell_b) for h in
           ["Empleado", "Área", "Días", "Tard.", "Faltas", "% Cumpl.",
            "Incidencias del período"]]
    filas = [enc]
    for e in empleados:
        inc = [f"{d.day:02d}&nbsp;{code}" for d, code in zip(dias, e["codigos"])
               if code in ("T", "F", "I")]
        inc_txt = " · ".join(inc) if inc else "Sin incidencias"
        filas.append([
            Paragraph(e["nombre"], cell),
            Paragraph(e["area"] or "—", cell),
            str(e["dias_trab"]), str(e["tardanzas"]), str(e["faltas"]),
            f"{e['porcentaje']}%",
            Paragraph(inc_txt, cell),
        ])
    w = doc.width
    t2 = Table(filas, repeatRows=1,
               colWidths=[w*0.17, w*0.12, w*0.06, w*0.06, w*0.06, w*0.08, w*0.45])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16A34A")),
        ("FONTSIZE",   (0, 0), (-1, -1), 7.5),
        ("ALIGN",      (2, 0), (5, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0FDF4")]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#D1FAE5")),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ]))
    story.append(t2)

    # ── Hojas-perfil por empleado (solo informe mensual) ──
    if detalle_por_empleado:
        for e in empleados:
            story.append(PageBreak())
            story.extend(_hoja_perfil(e, dias, doc, styles, colors, Table,
                                      TableStyle, Paragraph, Spacer,
                                      ParagraphStyle, cm))

    doc.build(story)
    buf.seek(0)
    return buf


def _hoja_perfil(e: dict, dias: list[date], doc, styles, colors, Table,
                 TableStyle, Paragraph, Spacer, ParagraphStyle, cm) -> list:
    """Una página tipo 'perfil' del empleado: cabecera + KPIs + gráfico de
    tardanzas + detalle de incidencias y justificaciones del mes."""
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart

    cell = ParagraphStyle("hp", parent=styles["Normal"], fontSize=8.5, leading=11)
    flow = []

    # Cabecera
    flow.append(Paragraph(f"<b>{e['nombre']}</b>", styles["Heading2"]))
    flow.append(Paragraph(
        f"Cargo: {e['cargo'] or '—'} &nbsp;·&nbsp; Área: {e['area'] or '—'}", cell))
    flow.append(Paragraph(f"Turno: {e['turno_label']}", cell))
    flow.append(Spacer(1, 0.3*cm))

    # KPIs en tarjetas (tabla de 6)
    kpis = [
        ("Días trabajados", str(e["dias_trab"])),
        ("Puntualidad",     f"{e['pct_punt']}%"),
        ("Prom. tardanza",  f"{e['prom_tarde']} min"),
        ("Tardanzas",       str(e["tardanzas"])),
        ("Faltas",          str(e["faltas"])),
        ("Horas (real/esp.)", f"{e['horas_reales']} / {e['horas_esper']}"),
    ]
    kp_lbl = [Paragraph(f"<b>{v}</b>", ParagraphStyle(
        "kv", parent=cell, fontSize=12, alignment=1)) for _, v in kpis]
    kp_cap = [Paragraph(k, ParagraphStyle(
        "kc", parent=cell, fontSize=7.5, alignment=1,
        textColor=colors.HexColor("#6B7280"))) for k, _ in kpis]
    cw = doc.width / 6
    tk = Table([kp_lbl, kp_cap], colWidths=[cw]*6)
    tk.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F0FDF4")),
        ("BOX",        (0, 0), (-1, -1), 0.5, colors.HexColor("#86EFAC")),
        ("INNERGRID",  (0, 0), (-1, -1), 0.5, colors.white),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 6),
    ]))
    flow.append(tk)
    flow.append(Spacer(1, 0.4*cm))

    # Gráfico de tardanzas por día (solo días con tardanza)
    puntos = [(d.day, m) for d, m in zip(dias, e["mins_tarde"]) if m > 0]
    if puntos:
        flow.append(Paragraph("<b>Minutos de tardanza por día</b>", styles["Heading4"]))
        dr = Drawing(doc.width, 4.2*cm)
        bc = VerticalBarChart()
        bc.x = 25; bc.y = 12
        bc.width = doc.width - 50; bc.height = 3.4*cm
        bc.data = [[m for _, m in puntos]]
        bc.categoryAxis.categoryNames = [str(d) for d, _ in puntos]
        bc.categoryAxis.labels.fontSize = 7
        bc.valueAxis.valueMin = 0
        bc.valueAxis.labels.fontSize = 7
        bc.bars[0].fillColor = colors.HexColor("#F59E0B")
        bc.barWidth = 8
        dr.add(bc)
        flow.append(dr)
        flow.append(Spacer(1, 0.3*cm))

    # Detalle de incidencias del mes
    flow.append(Paragraph("<b>Incidencias del mes</b>", styles["Heading4"]))
    label = {"T": "Tardanza", "F": "Falta", "I": "Incompleto"}
    filas_i = [[Paragraph("<b>Fecha</b>", cell), Paragraph("<b>Tipo</b>", cell),
                Paragraph("<b>Detalle</b>", cell)]]
    for d, code, mt in zip(dias, e["codigos"], e["mins_tarde"]):
        if code in label:
            det = f"{mt} min tarde" if code == "T" else "—"
            filas_i.append([d.strftime("%d/%m/%Y"), label[code], det])
    if len(filas_i) == 1:
        filas_i.append([Paragraph("<i>Sin incidencias en el período.</i>", cell), "", ""])
    ti = Table(filas_i, colWidths=[doc.width*0.2, doc.width*0.25, doc.width*0.55])
    ti.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FEF3C7")),
        ("FONTSIZE",   (0, 0), (-1, -1), 8.5),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    flow.append(ti)

    # Justificaciones aprobadas
    flow.append(Spacer(1, 0.35*cm))
    flow.append(Paragraph("<b>Justificaciones / permisos aprobados</b>", styles["Heading4"]))
    if e["justificaciones"]:
        filas_j = [[Paragraph("<b>Tipo</b>", cell), Paragraph("<b>Período</b>", cell),
                    Paragraph("<b>Descripción</b>", cell)]]
        for j in e["justificaciones"]:
            rango = (f"{j['desde'].strftime('%d/%m/%Y')} al {j['hasta'].strftime('%d/%m/%Y')}"
                     if j["desde"] and j["hasta"] else "—")
            filas_j.append([j["tipo"], rango, Paragraph(j["desc"] or "—", cell)])
        tj = Table(filas_j, colWidths=[doc.width*0.22, doc.width*0.28, doc.width*0.50])
        tj.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DBEAFE")),
            ("FONTSIZE",   (0, 0), (-1, -1), 8.5),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        flow.append(tj)
    else:
        flow.append(Paragraph("<i>Sin justificaciones registradas en el período.</i>", cell))

    return flow


def _export_response(buf: io.BytesIO, fmt: str, filename: str):
    if fmt == "pdf":
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}.pdf"'},
        )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


# ── Reporte Global (todos los empleados, rango libre) ────────────────────────

@router.get("/rrhh/asistencia/reportes/global")
def reporte_global(
    fecha_inicio: date = Query(...),
    fecha_fin:    date = Query(...),
    fmt:          str  = Query("xlsx", pattern="^(xlsx|pdf)$"),
    db:           Session = Depends(get_db),
    payload:      dict    = Depends(verificar_token),
):
    exigir_no_tecnico(payload)
    empresa_id = payload["empresa_id"]

    if fecha_fin < fecha_inicio:
        raise HTTPException(400, "fecha_fin debe ser >= fecha_inicio")

    filas, meta = _build_resumen_full(db, empresa_id, fecha_inicio, fecha_fin)
    titulo    = f"Reporte Global — {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"
    subtitulo = f"Meta: {meta:.0f} horas | Total empleados: {len(filas)}"
    encabezados = ["Empleado", "Cargo", "Área",
                   "H. Reales", "H. Justificadas", "H. Total",
                   "H. Faltantes", "% Cumpl.", "Faltas", "Tardanzas",
                   "Prom. Diario (h)"]
    dias_lab_count = len(_dias_laborables(fecha_inicio, fecha_fin)) or 1
    datos = [
        [r["nombre"], r["cargo"], r["area"],
         r["horas_reales"], r["horas_justificadas"], r["horas_total"],
         r["horas_faltantes"], f"{r['porcentaje']}%", r["faltas"], r["tardanzas"],
         round(r["horas_total"] / dias_lab_count, 2)]
        for r in filas
    ]

    if fmt == "xlsx":
        buf = _xlsx_stream(titulo, encabezados, datos)
    elif (fecha_fin - fecha_inicio).days <= 31:
        # Rango acotado → cronograma visual. Rango largo → tabla resumen (el
        # cronograma tendría demasiadas columnas para una página).
        dias_c, emps_c, tot_c = _build_cronograma(db, empresa_id, fecha_inicio, fecha_fin)
        buf = _pdf_cronograma(titulo, subtitulo, dias_c, emps_c, tot_c)
    else:
        buf = _pdf_stream(titulo, encabezados, datos, subtitulo)
    return _export_response(buf, fmt, f"reporte_global_{fecha_inicio.isoformat()}")


# ── Reporte Semanal ───────────────────────────────────────────────────────────

@router.get("/rrhh/asistencia/reportes/semanal")
def reporte_semanal(
    fecha_inicio: Optional[date] = Query(None),
    fecha_fin:    Optional[date] = Query(None),
    fmt:          str            = Query("xlsx", pattern="^(xlsx|pdf)$"),
    db:           Session        = Depends(get_db),
    payload:      dict           = Depends(verificar_token),
):
    exigir_no_tecnico(payload)
    empresa_id = payload["empresa_id"]

    hoy    = date.today()
    inicio = fecha_inicio or _lunes_semana(hoy)
    fin    = fecha_fin    or _sabado_semana(hoy)

    filas, meta = _build_resumen_full(db, empresa_id, inicio, fin)
    titulo    = f"Asistencia Semanal — {inicio.strftime('%d/%m/%Y')} al {fin.strftime('%d/%m/%Y')}"
    subtitulo = f"Meta: {meta} horas"
    encabezados = ["Empleado", "Cargo", "Área", "H. Reales", "H. Justificadas",
                   "H. Total", "H. Faltantes", "% Cumplimiento", "Faltas", "Tardanzas"]
    datos = [
        [r["nombre"], r["cargo"], r["area"],
         r["horas_reales"], r["horas_justificadas"], r["horas_total"],
         r["horas_faltantes"], f"{r['porcentaje']}%", r["faltas"], r["tardanzas"]]
        for r in filas
    ]

    if fmt == "xlsx":
        buf = _xlsx_stream(titulo, encabezados, datos)
    else:
        dias_c, emps_c, tot_c = _build_cronograma(db, empresa_id, inicio, fin)
        buf = _pdf_cronograma(
            titulo, f"Período: {inicio.strftime('%d/%m/%Y')} al {fin.strftime('%d/%m/%Y')}",
            dias_c, emps_c, tot_c)

    return _export_response(buf, fmt, f"asistencia_semanal_{inicio.isoformat()}")


# ── Excel Gerencial (Reporte Mensual Rico) ────────────────────────────────────

def _xlsx_mensual_gerencial(
    db: Session, empresa_id: str, inicio: date, fin: date, mes_label: str
) -> io.BytesIO:
    """
    Excel de gerencia con 5 hojas:
      1. Resumen Ejecutivo  — KPIs + tabla semáforo + gráfico de barras
      2. Cronograma         — Matriz día×empleado con colores semáforo
      3. Incidencias        — Faltas, tardanzas e incompletos por día con minutos tarde
      4. Horas Extra        — Empleados que superaron la meta mensual + gráfico
      5. Justificaciones    — Permisos/licencias aprobados en el período
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference

    # ── Paleta de colores (hex sin #) ──────────────────────────────────────────
    CV_TITULO   = "15803D"   # verde oscuro — encabezados principales
    CV_HEADER   = "BBF7D0"   # verde claro — encabezados de columna
    CV_CELDA    = "DCFCE7"   # semáforo: puntual
    CA_CELDA    = "FEF3C7"   # semáforo: tardanza
    CR_CELDA    = "FEE2E2"   # semáforo: falta
    CZ_CELDA    = "DBEAFE"   # semáforo: justificado / horas extra
    CO_CELDA    = "FFEDD5"   # semáforo: incompleto
    CG_CELDA    = "F3F4F6"   # semáforo: descanso
    BLANCO      = "FFFFFF"
    GRIS_TEXTO  = "64748B"

    CRONO_BG   = {"P": CV_CELDA, "T": CA_CELDA, "F": CR_CELDA,
                  "I": CO_CELDA, "J": CZ_CELDA,  "D": CG_CELDA, "H": "EDE9FE"}
    CRONO_FG   = {"P": "15803D", "T": "92400E", "F": "991B1B",
                  "I": "9A3412", "J": "1E40AF",  "D": "6B7280", "H": "6D28D9"}

    def _pct_color(pct: float) -> str:
        return CV_CELDA if pct >= 95 else CA_CELDA if pct >= 70 else CR_CELDA

    def _thin():
        s = Side(style="thin", color="D1FAE5")
        return Border(left=s, right=s, top=s, bottom=s)

    def _hdr(ws, row, col, val, bg=None, fg="FFFFFF", size=10, wrap=False, colspan=1):
        if colspan > 1:
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=col + colspan - 1)
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, size=size, color=fg)
        c.fill = PatternFill("solid", fgColor=bg or CV_TITULO)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        c.border = _thin()
        return c

    def _cel(ws, row, col, val, bg=None, bold=False, align="center", wrap=False, fmt=None):
        c = ws.cell(row=row, column=col, value=val)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(bold=bold, size=9)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        c.border = _thin()
        if fmt:
            c.number_format = fmt
        return c

    # ── Datos ──────────────────────────────────────────────────────────────────
    dias_all, emps_data, totales = _build_cronograma(db, empresa_id, inicio, fin)

    # Horas extra reales = reales − esperadas (por mes, no por semana)
    for emp in emps_data:
        emp["horas_extra"] = max(0.0, round(emp["horas_reales"] - emp["horas_esper"], 1))

    # Mapa empleado_nombre → id para buscar registros
    emp_id_map: dict[str, str] = {
        f"{u.nombre} {u.apellido}".strip(): e.id
        for e, u in (
            db.query(Empleado, Usuario)
            .join(Usuario, Usuario.id == Empleado.usuario_id)
            .filter(Empleado.empresa_id == empresa_id, Empleado.activo == True)
            .all()
        )
    }

    # Todos los registros del mes indexados por (empleado_id, fecha)
    ini_dt = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)
    fin_dt = datetime(fin.year,    fin.month,    fin.day,    23, 59, 59)
    todos_regs = (
        db.query(RegistroAsistencia)
        .filter(
            RegistroAsistencia.empresa_id == empresa_id,
            RegistroAsistencia.fecha_hora.between(ini_dt, fin_dt),
        ).all()
    )
    regs_x_dia: dict[tuple, list] = {}
    for r in todos_regs:
        regs_x_dia.setdefault((r.empleado_id, r.fecha_hora.date()), []).append(r)

    periodo_label = f"{inicio.strftime('%d/%m/%Y')} al {fin.strftime('%d/%m/%Y')}"
    n_emp = totales["n_emp"]
    pct_g = totales["pct_global"]
    tot_f = totales["faltas"]
    tot_t = totales["tardanzas"]
    tot_hx = sum(e["horas_extra"] for e in emps_data)

    # Distribución de estados para gráfico de torta
    dist: dict[str, int] = {"P": 0, "T": 0, "F": 0, "I": 0, "J": 0, "H": 0}
    for emp in emps_data:
        for code in emp["codigos"]:
            if code in dist:
                dist[code] += 1

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 1: RESUMEN EJECUTIVO
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = "15803D"

    # Título
    ws1.merge_cells("A1:J1")
    c = ws1.cell(row=1, column=1,
                 value=f"REPORTE DE ASISTENCIA  —  {mes_label.upper()}")
    c.font = Font(bold=True, size=16, color=BLANCO)
    c.fill = PatternFill("solid", fgColor=CV_TITULO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:J2")
    c2 = ws1.cell(row=2, column=1, value=f"Período: {periodo_label}")
    c2.font = Font(size=10, color=GRIS_TEXTO, italic=True)
    c2.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 20

    # KPI cards (fila 4-5)
    ws1.row_dimensions[3].height = 8
    ws1.row_dimensions[4].height = 22
    ws1.row_dimensions[5].height = 30
    ws1.row_dimensions[6].height = 8

    kpis = [
        ("Total Personal",   str(n_emp),          "1F2937", "F8FAFC"),
        ("Asistencia Global", f"{pct_g}%",
         CV_TITULO if pct_g >= 95 else "B45309" if pct_g >= 70 else "B91C1C",
         _pct_color(pct_g)),
        ("Total Faltas",     str(tot_f),           "B91C1C", CR_CELDA),
        ("Total Tardanzas",  str(tot_t),           "92400E", CA_CELDA),
        ("Horas Extra",      f"{tot_hx:.1f}h",     "1D4ED8", CZ_CELDA),
    ]
    for idx, (label, val, fcolor, bg) in enumerate(kpis):
        sc = idx * 2 + 1
        ws1.merge_cells(start_row=4, start_column=sc, end_row=4, end_column=sc + 1)
        c = ws1.cell(row=4, column=sc, value=label)
        c.font = Font(bold=True, size=9, color=GRIS_TEXTO)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws1.merge_cells(start_row=5, start_column=sc, end_row=5, end_column=sc + 1)
        c2 = ws1.cell(row=5, column=sc, value=val)
        c2.font = Font(bold=True, size=22, color=fcolor)
        c2.fill = PatternFill("solid", fgColor=bg)
        c2.alignment = Alignment(horizontal="center", vertical="center")

    # Encabezados tabla
    hdrs1 = ["N°", "Empleado", "Cargo", "Área",
             "Días Trab.", "Faltas", "Tardanzas", "H. Reales", "H. Extra", "% Cumpl."]
    R_TABLA = 8
    ws1.row_dimensions[R_TABLA].height = 22
    for ci, h in enumerate(hdrs1, 1):
        _hdr(ws1, R_TABLA, ci, h)

    emps_sorted = sorted(emps_data, key=lambda e: e["porcentaje"])
    for i, emp in enumerate(emps_sorted, 1):
        ri = R_TABLA + i
        ws1.row_dimensions[ri].height = 18
        row_bg = BLANCO if i % 2 == 0 else "F0FDF4"
        _cel(ws1, ri, 1, i,              bg=row_bg)
        _cel(ws1, ri, 2, emp["nombre"],  bg=row_bg, align="left")
        _cel(ws1, ri, 3, emp["cargo"],   bg=row_bg, align="left")
        _cel(ws1, ri, 4, emp["area"] or "—", bg=row_bg, align="left")
        _cel(ws1, ri, 5, emp["dias_trab"], bg=row_bg)
        _cel(ws1, ri, 6, emp["faltas"],
             bg=CR_CELDA if emp["faltas"] > 0 else row_bg,
             bold=emp["faltas"] > 0)
        _cel(ws1, ri, 7, emp["tardanzas"],
             bg=CA_CELDA if emp["tardanzas"] > 0 else row_bg,
             bold=emp["tardanzas"] > 0)
        _cel(ws1, ri, 8, emp["horas_reales"], bg=row_bg)
        _cel(ws1, ri, 9, emp["horas_extra"],
             bg=CZ_CELDA if emp["horas_extra"] > 0 else row_bg,
             bold=emp["horas_extra"] > 0)
        pct_num = emp["porcentaje"]
        _cel(ws1, ri, 10, pct_num / 100,
             bg=_pct_color(pct_num), bold=True, fmt='0.0%')

    # Fila totales
    TR = R_TABLA + n_emp + 1
    ws1.row_dimensions[TR].height = 22
    ws1.merge_cells(start_row=TR, start_column=1, end_row=TR, end_column=4)
    c = ws1.cell(row=TR, column=1, value="TOTALES / PROMEDIO")
    c.font = Font(bold=True, color=BLANCO); c.fill = PatternFill("solid", fgColor=CV_TITULO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    _cel(ws1, TR, 5,  sum(e["dias_trab"]   for e in emps_data), bg=CV_HEADER, bold=True)
    _cel(ws1, TR, 6,  tot_f,
         bg=CR_CELDA if tot_f > 0 else CV_HEADER, bold=True)
    _cel(ws1, TR, 7,  tot_t,
         bg=CA_CELDA if tot_t > 0 else CV_HEADER, bold=True)
    _cel(ws1, TR, 8,  round(sum(e["horas_reales"] for e in emps_data), 1),
         bg=CV_HEADER, bold=True)
    _cel(ws1, TR, 9,  round(tot_hx, 1),
         bg=CZ_CELDA if tot_hx > 0 else CV_HEADER, bold=True)
    _cel(ws1, TR, 10, pct_g / 100, bg=_pct_color(pct_g), bold=True, fmt='0.0%')

    # Leyenda semáforo
    LR = TR + 2
    ws1.merge_cells(start_row=LR, start_column=1, end_row=LR, end_column=10)
    lc = ws1.cell(row=LR, column=1,
        value="SEMÁFORO:  ■ Verde ≥ 95% cumplimiento   ■ Amarillo 70–94%   ■ Rojo < 70%")
    lc.font = Font(italic=True, size=9, color=GRIS_TEXTO)
    lc.alignment = Alignment(horizontal="left")

    # Anchos col hoja 1
    for ci, w in enumerate([5, 26, 18, 14, 10, 8, 10, 10, 9, 9], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # Gráfico barras: % por empleado
    CHR1 = LR + 3
    ch1 = BarChart()
    ch1.type = "col"; ch1.grouping = "clustered"
    ch1.title = f"% Cumplimiento por Empleado — {mes_label}"
    ch1.y_axis.title = "Cumplimiento"; ch1.y_axis.numFmt = '0%'
    ch1.y_axis.scaling.min = 0; ch1.y_axis.scaling.max = 1.1
    ch1.width = 28; ch1.height = 14; ch1.style = 10
    d_ref = Reference(ws1, min_col=10, min_row=R_TABLA,
                      max_row=R_TABLA + n_emp)
    c_ref = Reference(ws1, min_col=2,  min_row=R_TABLA + 1,
                      max_row=R_TABLA + n_emp)
    ch1.add_data(d_ref, titles_from_data=True)
    ch1.set_categories(c_ref)
    ws1.add_chart(ch1, f"A{CHR1}")

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 2: CRONOGRAMA
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Cronograma")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "0D9488"

    n_dias = len(dias_all)
    N_COLS2 = n_dias + 2 + 6   # nombre, cargo, días, + 6 KPI cols

    ws2.merge_cells(start_row=1, start_column=1,
                    end_row=1, end_column=N_COLS2)
    c = ws2.cell(row=1, column=1,
                 value=f"CRONOGRAMA DE ASISTENCIA — {mes_label.upper()}")
    c.font = Font(bold=True, size=14, color=BLANCO)
    c.fill = PatternFill("solid", fgColor="0D9488")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    ws2.row_dimensions[2].height = 38
    _hdr(ws2, 2, 1, "Empleado", wrap=True)
    _hdr(ws2, 2, 2, "Cargo",    wrap=True)

    for ci, dia in enumerate(dias_all, start=3):
        dom = dia.weekday() == 6
        bg  = "9CA3AF" if dom else "0D9488"
        c   = ws2.cell(row=2, column=ci,
                       value=f"{dia.day}\n{_DIAS_ES[dia.weekday()][:2]}")
        c.font      = Font(bold=True, size=7, color=BLANCO)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _thin()

    kpi_labels2 = ["T.Días", "Faltas", "Tard.", "Incompl.", "H. Real", "% Cumpl."]
    for i, lab in enumerate(kpi_labels2, start=n_dias + 3):
        _hdr(ws2, 2, i, lab, wrap=True)

    for ri, emp in enumerate(emps_data, start=3):
        ws2.row_dimensions[ri].height = 15
        _cel(ws2, ri, 1, emp["nombre"], align="left")
        _cel(ws2, ri, 2, emp["cargo"],  align="left")
        for ci, code in enumerate(emp["codigos"], start=3):
            c = ws2.cell(row=ri, column=ci, value=code)
            c.fill      = PatternFill("solid", fgColor=CRONO_BG.get(code, BLANCO))
            c.font      = Font(bold=True, size=8, color=CRONO_FG.get(code, "1F2937"))
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = _thin()
        ks = n_dias + 3
        _cel(ws2, ri, ks,   emp["dias_trab"])
        _cel(ws2, ri, ks+1, emp["faltas"],
             bg=CR_CELDA if emp["faltas"] > 0 else None, bold=emp["faltas"] > 0)
        _cel(ws2, ri, ks+2, emp["tardanzas"],
             bg=CA_CELDA if emp["tardanzas"] > 0 else None, bold=emp["tardanzas"] > 0)
        _cel(ws2, ri, ks+3, emp["incompletos"])
        _cel(ws2, ri, ks+4, emp["horas_reales"])
        _cel(ws2, ri, ks+5, emp["porcentaje"] / 100,
             bg=_pct_color(emp["porcentaje"]), bold=True, fmt='0%')

    # Leyenda cronograma
    LR2 = n_emp + 5
    ws2.merge_cells(start_row=LR2, start_column=1, end_row=LR2, end_column=2)
    ws2.cell(row=LR2, column=1, value="LEYENDA:").font = Font(bold=True, size=9)
    leyenda_items = [("P","Puntual"),("T","Tardanza"),("F","Falta"),
                     ("I","Incompl."),("J","Justificado"),("D","Descanso"),("H","Feriado")]
    for idx, (code, label) in enumerate(leyenda_items, start=3):
        c = ws2.cell(row=LR2, column=idx, value=f"■ {label}")
        c.font = Font(bold=True, size=8, color=CRONO_FG[code])
        c.fill = PatternFill("solid", fgColor=CRONO_BG[code])
        c.alignment = Alignment(horizontal="center")

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 14
    for ci in range(3, n_dias + 3):
        ws2.column_dimensions[get_column_letter(ci)].width = 3.8
    for ci in range(n_dias + 3, n_dias + 9):
        ws2.column_dimensions[get_column_letter(ci)].width = 8

    # Gráfico torta: distribución de estados
    # Escribimos la data en celdas ocultas para la referencia del gráfico
    DIST_ROW = LR2 + 2
    dist_labels = [("Puntual", dist["P"]), ("Tardanza", dist["T"]),
                   ("Falta",   dist["F"]), ("Incompl.", dist["I"]),
                   ("Justif.", dist["J"]), ("Feriado",  dist["H"])]
    for i, (lab, val) in enumerate(dist_labels):
        ws2.cell(row=DIST_ROW,   column=i+1, value=lab)
        ws2.cell(row=DIST_ROW+1, column=i+1, value=val)

    pie = PieChart()
    pie.title  = f"Distribución de Asistencia — {mes_label}"
    pie.width  = 16; pie.height = 10; pie.style = 10
    d_ref2 = Reference(ws2, min_col=1, max_col=6, min_row=DIST_ROW+1)
    l_ref2 = Reference(ws2, min_col=1, max_col=6, min_row=DIST_ROW)
    pie.add_data(d_ref2)
    pie.set_categories(l_ref2)
    ws2.add_chart(pie, f"A{DIST_ROW + 3}")

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 3: INCIDENCIAS (Faltas, Tardanzas, Incompletos)
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Incidencias")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = "B91C1C"

    ws3.merge_cells("A1:H1")
    c = ws3.cell(row=1, column=1,
                 value=f"FALTAS, TARDANZAS E INCOMPLETOS — {mes_label.upper()}")
    c.font = Font(bold=True, size=14, color=BLANCO)
    c.fill = PatternFill("solid", fgColor="B91C1C")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 30

    hdrs3 = ["Empleado","Cargo","Área","Fecha","Día","Estado","Min. Tarde","Observación"]
    ws3.row_dimensions[2].height = 22
    for ci, h in enumerate(hdrs3, 1):
        c = ws3.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, size=10, color=BLANCO)
        c.fill = PatternFill("solid", fgColor="B91C1C")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin()

    ri3 = 3
    for emp in sorted(emps_data, key=lambda e: e["nombre"]):
        emp_id = emp_id_map.get(emp["nombre"], "")
        for idx, (dia, code) in enumerate(zip(dias_all, emp["codigos"])):
            if code not in ("F", "T", "I"):
                continue
            regs_dia = regs_x_dia.get((emp_id, dia), [])
            entrada_r = next((r for r in regs_dia if r.tipo == "entrada"), None)
            salida_r  = next((r for r in regs_dia if r.tipo == "salida"),  None)

            estado_txt = {"F": "Falta", "T": "Tardanza", "I": "Incompleto"}[code]
            min_tarde  = emp["mins_tarde"][idx] if code == "T" else 0

            obs_parts = []
            if entrada_r:
                obs_parts.append(f"Entrada: {entrada_r.fecha_hora.strftime('%H:%M')}")
            if salida_r:
                obs_parts.append(f"Salida: {salida_r.fecha_hora.strftime('%H:%M')}")
            if code == "F":
                obs_parts.append("Sin registro de entrada")
            elif code == "I" and not salida_r:
                obs_parts.append("Sin registro de salida")
            obs = "  ·  ".join(obs_parts) or "—"

            bg3 = {"F": CR_CELDA, "T": CA_CELDA, "I": CO_CELDA}[code]
            ws3.row_dimensions[ri3].height = 18
            _cel(ws3, ri3, 1, emp["nombre"],        bg=bg3, align="left")
            _cel(ws3, ri3, 2, emp["cargo"],          bg=bg3, align="left")
            _cel(ws3, ri3, 3, emp["area"] or "—",   bg=bg3, align="left")
            _cel(ws3, ri3, 4, dia.strftime("%d/%m/%Y"), bg=bg3)
            _cel(ws3, ri3, 5, _DIAS_ES[dia.weekday()],  bg=bg3)
            _cel(ws3, ri3, 6, estado_txt,            bg=bg3, bold=True)
            _cel(ws3, ri3, 7,
                 f"{min_tarde} min" if min_tarde > 0 else "—",
                 bg=bg3 if min_tarde > 0 else None)
            _cel(ws3, ri3, 8, obs, align="left", wrap=True)
            ri3 += 1

    if ri3 == 3:
        ws3.merge_cells("A3:H3")
        c = ws3.cell(row=3, column=1,
                     value="✓ Sin incidencias en el período")
        c.font = Font(italic=True, color=CV_TITULO)
        c.fill = PatternFill("solid", fgColor=CV_CELDA)
        c.alignment = Alignment(horizontal="center")

    for ci, w in enumerate([26, 18, 14, 12, 10, 12, 10, 42], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    # Gráfico barras: tardanzas + faltas por empleado
    if ri3 > 3:
        emp_names_chart = [e["nombre"] for e in emps_data if e["faltas"] + e["tardanzas"] > 0]
        if len(emp_names_chart) >= 2:
            # Escribimos data auxiliar para el gráfico
            AUX_ROW3 = ri3 + 2
            ws3.cell(row=AUX_ROW3, column=1, value="Empleado")
            ws3.cell(row=AUX_ROW3, column=2, value="Faltas")
            ws3.cell(row=AUX_ROW3, column=3, value="Tardanzas")
            for ii, e in enumerate([x for x in emps_data
                                    if x["faltas"] + x["tardanzas"] > 0], 1):
                ws3.cell(row=AUX_ROW3+ii, column=1, value=e["nombre"])
                ws3.cell(row=AUX_ROW3+ii, column=2, value=e["faltas"])
                ws3.cell(row=AUX_ROW3+ii, column=3, value=e["tardanzas"])
            n_aux3 = len(emp_names_chart)
            ch3 = BarChart()
            ch3.type = "bar"; ch3.grouping = "clustered"
            ch3.title = f"Faltas y Tardanzas por Empleado — {mes_label}"
            ch3.style = 10; ch3.width = 24; ch3.height = 12
            d3 = Reference(ws3, min_col=2, max_col=3,
                           min_row=AUX_ROW3, max_row=AUX_ROW3+n_aux3)
            c3 = Reference(ws3, min_col=1,
                           min_row=AUX_ROW3+1, max_row=AUX_ROW3+n_aux3)
            ch3.add_data(d3, titles_from_data=True)
            ch3.set_categories(c3)
            ws3.add_chart(ch3, f"A{AUX_ROW3 + n_aux3 + 3}")

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 4: HORAS EXTRA
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Horas Extra")
    ws4.sheet_view.showGridLines = False
    ws4.sheet_properties.tabColor = "1D4ED8"

    ws4.merge_cells("A1:F1")
    c = ws4.cell(row=1, column=1,
                 value=f"HORAS EXTRA — {mes_label.upper()}")
    c.font = Font(bold=True, size=14, color=BLANCO)
    c.fill = PatternFill("solid", fgColor="1D4ED8")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 30

    hdrs4 = ["Empleado","Cargo","Área","H. Reales","H. Esperadas","H. Extra"]
    ws4.row_dimensions[2].height = 22
    for ci, h in enumerate(hdrs4, 1):
        c = ws4.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, size=10, color=BLANCO)
        c.fill = PatternFill("solid", fgColor="1D4ED8")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin()

    emps_hx = sorted([e for e in emps_data if e["horas_extra"] > 0],
                     key=lambda e: -e["horas_extra"])
    for i, emp in enumerate(emps_hx, start=3):
        ws4.row_dimensions[i].height = 18
        bg4 = BLANCO if i % 2 else "EFF6FF"
        _cel(ws4, i, 1, emp["nombre"],      bg=bg4, align="left")
        _cel(ws4, i, 2, emp["cargo"],       bg=bg4, align="left")
        _cel(ws4, i, 3, emp["area"] or "—", bg=bg4, align="left")
        _cel(ws4, i, 4, emp["horas_reales"])
        _cel(ws4, i, 5, emp["horas_esper"])
        _cel(ws4, i, 6, emp["horas_extra"], bg=CZ_CELDA, bold=True)

    if not emps_hx:
        ws4.merge_cells("A3:F3")
        c = ws4.cell(row=3, column=1, value="Sin empleados con horas extra en el período")
        c.font = Font(italic=True, color=GRIS_TEXTO)
        c.alignment = Alignment(horizontal="center")

    for ci, w in enumerate([26, 18, 14, 11, 13, 11], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    if len(emps_hx) >= 2:
        R4C = len(emps_hx) + 5
        ch4 = BarChart()
        ch4.type = "col"; ch4.style = 10
        ch4.title = f"Horas Extra por Empleado — {mes_label}"
        ch4.y_axis.title = "Horas"; ch4.width = 22; ch4.height = 12
        d4 = Reference(ws4, min_col=6, min_row=2, max_row=2 + len(emps_hx))
        c4 = Reference(ws4, min_col=1, min_row=3, max_row=2 + len(emps_hx))
        ch4.add_data(d4, titles_from_data=True)
        ch4.set_categories(c4)
        ws4.add_chart(ch4, f"A{R4C}")

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 5: JUSTIFICACIONES
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Justificaciones")
    ws5.sheet_view.showGridLines = False
    ws5.sheet_properties.tabColor = "7C3AED"

    ws5.merge_cells("A1:G1")
    c = ws5.cell(row=1, column=1,
                 value=f"PERMISOS Y JUSTIFICACIONES APROBADAS — {mes_label.upper()}")
    c.font = Font(bold=True, size=14, color=BLANCO)
    c.fill = PatternFill("solid", fgColor="7C3AED")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 30

    hdrs5 = ["Empleado","Cargo","Área","Tipo","Desde","Hasta","Descripción"]
    ws5.row_dimensions[2].height = 22
    for ci, h in enumerate(hdrs5, 1):
        c = ws5.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, size=10, color=BLANCO)
        c.fill = PatternFill("solid", fgColor="7C3AED")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin()

    ri5 = 3
    for emp in sorted(emps_data, key=lambda e: e["nombre"]):
        for j in emp.get("justificaciones", []):
            ws5.row_dimensions[ri5].height = 18
            bg5 = BLANCO if ri5 % 2 else "F5F3FF"
            _cel(ws5, ri5, 1, emp["nombre"],      bg=bg5, align="left")
            _cel(ws5, ri5, 2, emp["cargo"],        bg=bg5, align="left")
            _cel(ws5, ri5, 3, emp["area"] or "—",  bg=bg5, align="left")
            _cel(ws5, ri5, 4, j.get("tipo", "Permiso"), bg=CZ_CELDA)
            desde = j.get("desde")
            hasta = j.get("hasta")
            _cel(ws5, ri5, 5,
                 desde.strftime("%d/%m/%Y") if isinstance(desde, date) else str(desde or "—"))
            _cel(ws5, ri5, 6,
                 hasta.strftime("%d/%m/%Y") if isinstance(hasta, date) else str(hasta or "—"))
            _cel(ws5, ri5, 7, j.get("desc", "") or "—",
                 align="left", wrap=True)
            ri5 += 1

    if ri5 == 3:
        ws5.merge_cells("A3:G3")
        c = ws5.cell(row=3, column=1, value="Sin justificaciones registradas en el período")
        c.font = Font(italic=True, color=GRIS_TEXTO)
        c.alignment = Alignment(horizontal="center")

    for ci, w in enumerate([26, 18, 14, 20, 12, 12, 44], 1):
        ws5.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Reporte Mensual ───────────────────────────────────────────────────────────

@router.get("/rrhh/asistencia/reportes/mensual")
def reporte_mensual(
    anio: int = Query(..., ge=2020, le=2100),
    mes:  int = Query(..., ge=1, le=12),
    fmt:  str = Query("xlsx", pattern="^(xlsx|pdf)$"),
    db:   Session = Depends(get_db),
    payload: dict = Depends(verificar_token),
):
    exigir_no_tecnico(payload)
    empresa_id = payload["empresa_id"]

    from calendar import monthrange
    inicio = date(anio, mes, 1)
    fin    = date(anio, mes, monthrange(anio, mes)[1])

    meses_es = ["", "Enero","Febrero","Marzo","Abril","Mayo","Junio",
                "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    mes_label = f"{meses_es[mes]} {anio}"
    titulo    = f"Asistencia Mensual — {mes_label}"

    if fmt == "xlsx":
        buf = _xlsx_mensual_gerencial(db, empresa_id, inicio, fin, mes_label)
    else:
        dias_c, emps_c, tot_c = _build_cronograma(db, empresa_id, inicio, fin)
        buf = _pdf_cronograma(
            titulo, f"Período: {inicio.strftime('%d/%m/%Y')} al {fin.strftime('%d/%m/%Y')}",
            dias_c, emps_c, tot_c, detalle_por_empleado=True)

    return _export_response(buf, fmt, f"asistencia_mensual_{anio}_{mes:02d}")


# ── Reporte Tardanzas/Faltas ──────────────────────────────────────────────────

@router.get("/rrhh/asistencia/reportes/tardanzas")
def reporte_tardanzas(
    fecha_inicio: date = Query(...),
    fecha_fin:    date = Query(...),
    fmt:          str  = Query("xlsx", pattern="^(xlsx|pdf)$"),
    db:           Session = Depends(get_db),
    payload:      dict    = Depends(verificar_token),
):
    exigir_no_tecnico(payload)
    empresa_id = payload["empresa_id"]

    filas, _ = _build_resumen_full(db, empresa_id, fecha_inicio, fecha_fin)
    titulo = f"Tardanzas y Faltas — {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"
    incidencias = [r for r in filas if r["faltas"] > 0 or r["tardanzas"] > 0]
    encabezados = ["Empleado", "Cargo", "Área", "Faltas", "Tardanzas", "H. Faltantes"]
    datos = [
        [r["nombre"], r["cargo"], r["area"], r["faltas"], r["tardanzas"], r["horas_faltantes"]]
        for r in sorted(incidencias, key=lambda x: -(x["faltas"] + x["tardanzas"]))
    ]

    if fmt == "xlsx":
        buf = _xlsx_stream(titulo, encabezados, datos)
    else:
        buf = _pdf_stream(titulo, encabezados, datos)

    return _export_response(buf, fmt, f"tardanzas_faltas_{fecha_inicio.isoformat()}")


# ── Reporte Individual ────────────────────────────────────────────────────────

@router.get("/rrhh/asistencia/reportes/individual")
def reporte_individual(
    empleado_id:  str  = Query(...),
    fecha_inicio: date = Query(...),
    fecha_fin:    date = Query(...),
    fmt:          str  = Query("xlsx", pattern="^(xlsx|pdf)$"),
    db:           Session = Depends(get_db),
    payload:      dict    = Depends(verificar_token),
):
    exigir_no_tecnico(payload)
    empresa_id = payload["empresa_id"]

    emp = db.query(Empleado).filter(
        Empleado.id == empleado_id, Empleado.empresa_id == empresa_id
    ).first()
    if not emp:
        raise HTTPException(404, "Empleado no encontrado")

    usr = db.query(Usuario).filter(Usuario.id == emp.usuario_id).first()
    nombre = f"{usr.nombre} {usr.apellido}".strip() if usr else "Empleado"

    dias_lab     = _dias_laborables(fecha_inicio, fecha_fin)
    inicio_dt    = datetime(fecha_inicio.year, fecha_inicio.month, fecha_inicio.day, 0, 0, 0)
    fin_dt       = datetime(fecha_fin.year,    fecha_fin.month,    fecha_fin.day,    23, 59, 59)

    registros = (
        db.query(RegistroAsistencia)
        .filter(
            RegistroAsistencia.empresa_id  == empresa_id,
            RegistroAsistencia.empleado_id == empleado_id,
            RegistroAsistencia.fecha_hora  >= inicio_dt,
            RegistroAsistencia.fecha_hora  <= fin_dt,
        )
        .all()
    )

    regs_por_dia: dict[date, list] = {}
    for r in registros:
        regs_por_dia.setdefault(r.fecha_hora.date(), []).append(r)

    titulo = f"Historial Asistencia — {nombre}"
    encabezados = ["Fecha", "Día", "Ingreso", "Salida", "H. Trabajo", "Estado"]
    datos = []
    for dia in sorted(dias_lab, reverse=True):
        regs_dia = regs_por_dia.get(dia, [])
        horas    = _horas_dia(regs_dia)
        estado   = _estado_dia(regs_dia, horas, True)
        entrada  = next((r.fecha_hora for r in regs_dia if r.tipo == "entrada"), None)
        salida   = next((r.fecha_hora for r in regs_dia if r.tipo == "salida"),  None)
        datos.append([
            dia.strftime("%d/%m/%Y"),
            _DIAS_ES[dia.weekday()],
            entrada.strftime("%H:%M") if entrada else "—",
            salida.strftime("%H:%M")  if salida  else "—",
            f"{horas:.1f}h",
            estado,
        ])

    if fmt == "xlsx":
        buf = _xlsx_stream(titulo, encabezados, datos)
    else:
        buf = _pdf_stream(titulo, encabezados, datos,
                          f"{fecha_inicio.strftime('%d/%m/%Y')} – {fecha_fin.strftime('%d/%m/%Y')}")

    return _export_response(buf, fmt, f"asistencia_{empleado_id[:8]}_{fecha_inicio.isoformat()}")


# ── Reporte Horas Extra ───────────────────────────────────────────────────────
# Basado en solicitudes de permanencia_extra APROBADAS con >= 1 hora (campo dias).

@router.get("/rrhh/asistencia/reportes/horas-extra")
def reporte_horas_extra(
    fecha_inicio: Optional[date] = Query(None),
    fecha_fin:    Optional[date] = Query(None),
    fmt:          str            = Query("xlsx", pattern="^(xlsx|pdf)$"),
    db:           Session        = Depends(get_db),
    payload:      dict           = Depends(verificar_token),
):
    exigir_no_tecnico(payload)
    empresa_id = payload["empresa_id"]

    hoy    = date.today()
    inicio = fecha_inicio or _lunes_semana(hoy)
    fin    = fecha_fin    or _sabado_semana(hoy)

    rows = (
        db.query(SolicitudLaboral, Empleado, Usuario)
        .join(Empleado, SolicitudLaboral.empleado_id == Empleado.id)
        .join(Usuario,  Empleado.usuario_id == Usuario.id)
        .filter(
            SolicitudLaboral.empresa_id    == empresa_id,
            SolicitudLaboral.tipo          == "permanencia_extra",
            SolicitudLaboral.estado        == "aprobada",
            SolicitudLaboral.dias          >= 1,
            SolicitudLaboral.fecha_inicio  >= inicio,
            SolicitudLaboral.fecha_inicio  <= fin,
        )
        .order_by(SolicitudLaboral.fecha_inicio)
        .all()
    )

    titulo = f"Horas Extra Aprobadas — {inicio.strftime('%d/%m/%Y')} al {fin.strftime('%d/%m/%Y')}"
    encabezados = ["Empleado", "Cargo", "Área", "Fecha", "Horas Aprobadas", "Ref. Solicitud", "Estado"]
    datos = [
        [
            f"{usr.nombre} {usr.apellido}".strip(),
            emp.cargo or "",
            emp.area  or "",
            sol.fecha_inicio.strftime("%d/%m/%Y"),
            sol.dias,
            str(sol.id)[:8].upper(),
            sol.estado.capitalize(),
        ]
        for sol, emp, usr in rows
    ]

    subtitulo = f"Solicitudes de Permanencia Extra aprobadas del {inicio.strftime('%d/%m/%Y')} al {fin.strftime('%d/%m/%Y')}"
    if fmt == "xlsx":
        buf = _xlsx_stream(titulo, encabezados, datos)
    else:
        buf = _pdf_stream(titulo, encabezados, datos, subtitulo)

    return _export_response(buf, fmt, f"horas_extra_{inicio.isoformat()}")

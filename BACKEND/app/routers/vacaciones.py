"""
Router: /vacaciones — Punto 3.3 (RR.HH.), control de vacaciones por ley (Perú).
Régimen parametrizable por empresa (General 30 d/año, REMYPE 15 d/año) con tope
de acumulación. Saldo derivado: devengo mensual desde fecha_ingreso menos los
días aprobados. Sin cálculo de pago vacacional.
Lecturas: empresa del token. Escrituras: RBAC dominio 'vacaciones'.
"""
from __future__ import annotations

import io
import uuid as _uuid
from datetime import date, datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..core.security import verificar_token
from ..core.permisos import exigir_permiso
from ..db.database import get_db
from ..models.empleado import Empleado
from ..models.usuario import Usuario
from ..models.vacaciones import AjusteSaldoVacaciones, ConfigVacaciones, SolicitudVacaciones
from ..models.solicitud_laboral import SolicitudLaboral
from ..services.fcm_service import notificar_usuario
from ..schemas.vacaciones import (
    AjusteSaldoIn, AjusteSaldoOut, ConfigIn, ConfigOut, SaldoOut, SolicitudIn, SolicitudOut,
    DetalleSolicitudVac,
)

router = APIRouter(prefix="/vacaciones", tags=["vacaciones"])

# Días/año por régimen (Perú). 'otro' usa el valor explícito enviado.
REGIMEN_DIAS = {"general": 30, "remype": 15}
TOPE_DEFECTO = 30


def _parse_date(s: str) -> date:
    return date.fromisoformat(s[:10])


def _meses_servicio(ingreso: Optional[date], hoy: date) -> int:
    if not ingreso:
        return 0
    meses = (hoy.year - ingreso.year) * 12 + (hoy.month - ingreso.month)
    if hoy.day < ingreso.day:
        meses -= 1
    return max(0, meses)


def _get_config(db: Session, empresa_id: str) -> ConfigVacaciones:
    cfg = db.query(ConfigVacaciones).filter(ConfigVacaciones.empresa_id == empresa_id).first()
    if not cfg:
        cfg = ConfigVacaciones(
            id=str(_uuid.uuid4()), empresa_id=empresa_id,
            regimen="general", dias_por_anio=REGIMEN_DIAS["general"], tope_acumulacion=TOPE_DEFECTO,
        )
        db.add(cfg)
        db.commit()
    return cfg


def _nombre(db: Session, empleado: Empleado) -> Optional[str]:
    row = db.query(Usuario.nombre, Usuario.apellido).filter(Usuario.id == empleado.usuario_id).first()
    return f"{row[0]} {row[1]}".strip() if row else None


def _gozado_detalle(db: Session, empresa_id: str, empleado_id: str) -> tuple[int, list]:
    """Suma los días de solicitudes_laborales de tipo vacaciones aprobadas."""
    rows = (
        db.query(SolicitudLaboral)
        .filter(
            SolicitudLaboral.empresa_id == empresa_id,
            SolicitudLaboral.empleado_id == empleado_id,
            SolicitudLaboral.tipo.ilike("vacaciones"),
            SolicitudLaboral.estado == "aprobada",
            SolicitudLaboral.fecha_inicio.isnot(None),
            SolicitudLaboral.fecha_fin.isnot(None),
        )
        .order_by(SolicitudLaboral.fecha_inicio)
        .all()
    )
    detalle: list[DetalleSolicitudVac] = []
    total = 0
    for r in rows:
        dias = (r.fecha_fin - r.fecha_inicio).days + 1
        total += dias
        detalle.append(DetalleSolicitudVac(
            fecha_inicio=r.fecha_inicio.isoformat(),
            fecha_fin=r.fecha_fin.isoformat(),
            dias=dias,
            fecha_aprobacion=r.fecha_aprobacion.isoformat() if r.fecha_aprobacion else None,
        ))
    return total, detalle


def _ajuste(db: Session, empleado_id: str) -> int:
    row = db.query(AjusteSaldoVacaciones).filter(
        AjusteSaldoVacaciones.empleado_id == empleado_id).first()
    return row.ajuste_dias if row else 0


def _estado_vac(meses: int, disponible: float) -> str:
    if meses < 12:
        return "sin_derecho"
    if disponible <= 0:
        return "agotado"
    return "disponible"


def _saldo(db: Session, empresa_id: str, emp: Empleado, cfg: ConfigVacaciones) -> SaldoOut:
    hoy = date.today()
    meses = _meses_servicio(emp.fecha_ingreso, hoy)
    anos = meses // 12
    devengado = float(anos * cfg.dias_por_anio)
    ajuste_dias = _ajuste(db, str(emp.id))
    gozado, solicitudes_gozadas = _gozado_detalle(db, empresa_id, str(emp.id))
    acumulado = devengado + ajuste_dias - gozado
    disponible = round(min(acumulado, float(cfg.tope_acumulacion)), 2) if acumulado > 0 else 0.0
    return SaldoOut(
        empleado_id=str(emp.id), empleado_nombre=_nombre(db, emp),
        cargo=emp.cargo,
        fecha_ingreso=(str(emp.fecha_ingreso) if emp.fecha_ingreso else None),
        meses_servicio=meses, anos_servicio=anos, dias_por_anio=cfg.dias_por_anio,
        devengado=devengado, ajuste_dias=ajuste_dias, gozado=gozado,
        disponible=disponible, tope_acumulacion=cfg.tope_acumulacion,
        estado_vacaciones=_estado_vac(meses, disponible),
        solicitudes_gozadas=solicitudes_gozadas,
    )


def _empleado_actual(db: Session, payload: dict) -> Optional[Empleado]:
    return (
        db.query(Empleado)
        .filter(Empleado.usuario_id == payload.get("id"),
                Empleado.empresa_id == payload["empresa_id"])
        .first()
    )


# ── Configuración ────────────────────────────────────────────────────────────
@router.get("/config", response_model=ConfigOut)
def obtener_config(payload: dict = Depends(verificar_token), db: Session = Depends(get_db)):
    cfg = _get_config(db, payload["empresa_id"])
    return ConfigOut(regimen=cfg.regimen, dias_por_anio=cfg.dias_por_anio,
                     tope_acumulacion=cfg.tope_acumulacion)


@router.put("/config", response_model=ConfigOut)
def actualizar_config(body: ConfigIn, payload: dict = Depends(verificar_token), db: Session = Depends(get_db)):
    exigir_permiso(db, payload, "vacaciones", "configurar")
    cfg = _get_config(db, payload["empresa_id"])
    cfg.regimen = body.regimen
    # Días/año: explícito o derivado del régimen (general/remype). 'otro' exige valor.
    if body.dias_por_anio is not None:
        cfg.dias_por_anio = body.dias_por_anio
    elif body.regimen in REGIMEN_DIAS:
        cfg.dias_por_anio = REGIMEN_DIAS[body.regimen]
    elif body.regimen == "otro":
        raise HTTPException(status_code=422, detail="Régimen 'otro' requiere dias_por_anio")
    cfg.tope_acumulacion = body.tope_acumulacion if body.tope_acumulacion is not None else cfg.tope_acumulacion
    cfg.updated_at = datetime.utcnow()
    db.commit()
    return ConfigOut(regimen=cfg.regimen, dias_por_anio=cfg.dias_por_anio,
                     tope_acumulacion=cfg.tope_acumulacion)


# ── Saldos ───────────────────────────────────────────────────────────────────
@router.get("/saldos", response_model=List[SaldoOut])
def listar_saldos(payload: dict = Depends(verificar_token), db: Session = Depends(get_db)):
    empresa_id = payload["empresa_id"]
    cfg = _get_config(db, empresa_id)
    emps = (
        db.query(Empleado)
        .filter(Empleado.empresa_id == empresa_id, Empleado.activo == True)  # noqa: E712
        .all()
    )
    saldos = [_saldo(db, empresa_id, e, cfg) for e in emps]
    saldos.sort(key=lambda s: (s.empleado_nombre or ""))
    return saldos


@router.get("/saldo/{empleado_id}", response_model=SaldoOut)
def saldo_empleado(empleado_id: str, payload: dict = Depends(verificar_token), db: Session = Depends(get_db)):
    empresa_id = payload["empresa_id"]
    emp = db.query(Empleado).filter(
        Empleado.id == empleado_id, Empleado.empresa_id == empresa_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    return _saldo(db, empresa_id, emp, _get_config(db, empresa_id))


# ── Solicitudes ──────────────────────────────────────────────────────────────
def _sol_out(db: Session, s: SolicitudVacaciones) -> SolicitudOut:
    nombre = (
        db.query(Usuario.nombre, Usuario.apellido)
        .join(Empleado, Empleado.usuario_id == Usuario.id)
        .filter(Empleado.id == s.empleado_id).first()
    )
    return SolicitudOut(
        id=str(s.id), empleado_id=str(s.empleado_id),
        empleado_nombre=(f"{nombre[0]} {nombre[1]}".strip() if nombre else None),
        fecha_inicio=(str(s.fecha_inicio) if s.fecha_inicio else None),
        fecha_fin=(str(s.fecha_fin) if s.fecha_fin else None),
        dias=int(s.dias), estado=s.estado, motivo=s.motivo,
        fecha_resolucion=(s.fecha_resolucion.isoformat() if s.fecha_resolucion else None),
    )


@router.get("/solicitudes", response_model=List[SolicitudOut])
def listar_solicitudes(
    empleado_id: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    payload: dict = Depends(verificar_token), db: Session = Depends(get_db),
):
    qry = db.query(SolicitudVacaciones).filter(
        SolicitudVacaciones.empresa_id == payload["empresa_id"])
    if empleado_id:
        qry = qry.filter(SolicitudVacaciones.empleado_id == empleado_id)
    if estado:
        qry = qry.filter(SolicitudVacaciones.estado == estado)
    rows = qry.order_by(SolicitudVacaciones.created_at.desc()).limit(500).all()
    return [_sol_out(db, s) for s in rows]


@router.post("/solicitudes", response_model=SolicitudOut, status_code=201)
def crear_solicitud(body: SolicitudIn, payload: dict = Depends(verificar_token), db: Session = Depends(get_db)):
    empresa_id = payload["empresa_id"]
    self_emp = _empleado_actual(db, payload)
    # Auto-solicitud (propio): libre para cualquier empleado con ficha.
    # Solicitar por OTRO empleado: requiere permiso vacaciones:solicitar.
    if body.empleado_id and (not self_emp or body.empleado_id != str(self_emp.id)):
        exigir_permiso(db, payload, "vacaciones", "solicitar")
        emp = db.query(Empleado).filter(
            Empleado.id == body.empleado_id, Empleado.empresa_id == empresa_id).first()
    else:
        emp = self_emp
    if not emp:
        raise HTTPException(status_code=404, detail="Tu usuario no tiene ficha de empleado")
    ini = _parse_date(body.fecha_inicio)
    fin = _parse_date(body.fecha_fin)
    if fin < ini:
        raise HTTPException(status_code=422, detail="fecha_fin no puede ser anterior a fecha_inicio")
    dias = (fin - ini).days + 1
    saldo = _saldo(db, empresa_id, emp, _get_config(db, empresa_id))
    if dias > saldo.disponible:
        raise HTTPException(
            status_code=409,
            detail=f"Saldo insuficiente: solicitas {dias} d, disponible {saldo.disponible:.1f} d")
    s = SolicitudVacaciones(
        id=str(_uuid.uuid4()), empresa_id=empresa_id, empleado_id=str(emp.id),
        fecha_inicio=ini, fecha_fin=fin, dias=dias, estado="pendiente",
        motivo=body.motivo, solicitado_por_id=payload.get("id"),
    )
    db.add(s)
    db.commit()
    return _sol_out(db, s)


# ── Autoservicio del empleado (token) ────────────────────────────────────────
@router.get("/mi-saldo", response_model=SaldoOut)
def mi_saldo(payload: dict = Depends(verificar_token), db: Session = Depends(get_db)):
    empresa_id = payload["empresa_id"]
    emp = _empleado_actual(db, payload)
    if not emp:
        raise HTTPException(status_code=404, detail="Tu usuario no tiene ficha de empleado")
    return _saldo(db, empresa_id, emp, _get_config(db, empresa_id))


@router.get("/mis-solicitudes", response_model=List[SolicitudOut])
def mis_solicitudes(payload: dict = Depends(verificar_token), db: Session = Depends(get_db)):
    emp = _empleado_actual(db, payload)
    if not emp:
        return []
    rows = (
        db.query(SolicitudVacaciones)
        .filter(SolicitudVacaciones.empresa_id == payload["empresa_id"],
                SolicitudVacaciones.empleado_id == str(emp.id))
        .order_by(SolicitudVacaciones.created_at.desc())
        .limit(200).all()
    )
    return [_sol_out(db, s) for s in rows]


def _resolver(db: Session, payload: dict, sol_id: str, nuevo_estado: str) -> SolicitudOut:
    empresa_id = payload["empresa_id"]
    s = db.query(SolicitudVacaciones).filter(
        SolicitudVacaciones.id == sol_id, SolicitudVacaciones.empresa_id == empresa_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada")
    if s.estado != "pendiente":
        raise HTTPException(status_code=409, detail=f"La solicitud ya está {s.estado}")
    if nuevo_estado == "aprobada":
        # Validar saldo disponible antes de aprobar
        emp = db.query(Empleado).filter(Empleado.id == s.empleado_id).first()
        cfg = _get_config(db, empresa_id)
        saldo = _saldo(db, empresa_id, emp, cfg)
        if s.dias > saldo.disponible:
            raise HTTPException(
                status_code=409,
                detail=f"Saldo insuficiente: solicita {s.dias} d, disponible {saldo.disponible} d")
    s.estado = nuevo_estado
    s.resuelto_por_id = payload.get("id")
    s.fecha_resolucion = datetime.utcnow()
    db.commit()

    # Notificar al solicitante el resultado de su solicitud.
    try:
        emp = db.query(Empleado).filter(Empleado.id == s.empleado_id).first()
        if emp and emp.usuario_id:
            if nuevo_estado == "aprobada":
                titulo = "Vacaciones aprobadas"
                mensaje = (f"Tu solicitud de {s.dias} día(s) "
                           f"({s.fecha_inicio.isoformat()} al {s.fecha_fin.isoformat()}) fue aprobada.")
            else:
                titulo = "Vacaciones rechazadas"
                mensaje = (f"Tu solicitud de {s.dias} día(s) "
                           f"({s.fecha_inicio.isoformat()} al {s.fecha_fin.isoformat()}) fue rechazada.")
            notificar_usuario(
                db, empresa_id=empresa_id, usuario_id=emp.usuario_id,
                titulo=titulo, mensaje=mensaje,
                tipo="recordatorio", categoria="vacaciones",
                referencia_tabla=f"vacaciones:{s.id}",
            )
            db.commit()
    except Exception:
        db.rollback()

    return _sol_out(db, s)


@router.post("/solicitudes/{sol_id}/aprobar", response_model=SolicitudOut)
def aprobar(sol_id: str, payload: dict = Depends(verificar_token), db: Session = Depends(get_db)):
    exigir_permiso(db, payload, "vacaciones", "aprobar")
    return _resolver(db, payload, sol_id, "aprobada")


@router.post("/solicitudes/{sol_id}/rechazar", response_model=SolicitudOut)
def rechazar(sol_id: str, payload: dict = Depends(verificar_token), db: Session = Depends(get_db)):
    exigir_permiso(db, payload, "vacaciones", "rechazar")
    return _resolver(db, payload, sol_id, "rechazada")


# ── Ajuste de saldo inicial (migración) ──────────────────────────────────────
@router.put("/saldo-inicial/{empleado_id}", response_model=AjusteSaldoOut)
def fijar_saldo_inicial(
    empleado_id: str,
    body: AjusteSaldoIn,
    payload: dict = Depends(verificar_token),
    db: Session = Depends(get_db),
):
    """Fija el saldo de vacaciones disponibles de un empleado a `dias_disponibles`,
    calculando el ajuste necesario según el devengado actual (años completos).
    Requiere permiso vacaciones:configurar. Idempotente (upsert)."""
    exigir_permiso(db, payload, "vacaciones", "configurar")
    empresa_id = payload["empresa_id"]
    emp = db.query(Empleado).filter(
        Empleado.id == empleado_id, Empleado.empresa_id == empresa_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    cfg = _get_config(db, empresa_id)
    meses = _meses_servicio(emp.fecha_ingreso, date.today())
    anos = meses // 12
    devengado_actual = float(anos * cfg.dias_por_anio)
    # ajuste = diferencia entre lo que queremos que tenga y lo que ya tiene por devengo
    ajuste = int(round(body.dias_disponibles - devengado_actual))
    row = db.query(AjusteSaldoVacaciones).filter(
        AjusteSaldoVacaciones.empleado_id == empleado_id).first()
    if row:
        row.ajuste_dias = ajuste
        row.notas = body.notas
        row.creado_por_id = payload.get("id")
        row.updated_at = datetime.utcnow()
    else:
        row = AjusteSaldoVacaciones(
            id=str(_uuid.uuid4()), empresa_id=empresa_id,
            empleado_id=empleado_id, ajuste_dias=ajuste,
            notas=body.notas, creado_por_id=payload.get("id"),
        )
        db.add(row)
    db.commit()
    return AjusteSaldoOut(
        empleado_id=empleado_id, ajuste_dias=ajuste, notas=row.notas)


# ── Helpers de reporte ────────────────────────────────────────────────────────

def _saldos_todos(db: Session, empresa_id: str):
    cfg = _get_config(db, empresa_id)
    emps = (
        db.query(Empleado)
        .filter(Empleado.empresa_id == empresa_id, Empleado.activo == True)  # noqa: E712
        .all()
    )
    saldos = [_saldo(db, empresa_id, e, cfg) for e in emps]
    saldos.sort(key=lambda s: (s.empleado_nombre or ""))
    return saldos, cfg


_ESTADO_LABEL = {
    "sin_derecho": "Sin derecho (<1 año)",
    "agotado":     "Saldo agotado",
    "disponible":  "Con días disponibles",
}


# ── Reporte Excel ─────────────────────────────────────────────────────────────

@router.get("/reporte.xlsx")
def reporte_excel(payload: dict = Depends(verificar_token), db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    empresa_id = payload["empresa_id"]
    saldos, cfg = _saldos_todos(db, empresa_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vacaciones"

    # ── Paleta ────────────────────────────────────────────────────────────────
    AZUL      = "1A56DB"
    BLANCO    = "FFFFFF"
    ROJO_BG   = "FEE2E2"
    ROJO_FG   = "991B1B"
    VERDE_BG  = "DCFCE7"
    VERDE_FG  = "166534"
    AMBAR_BG  = "FEF3C7"
    AMBAR_FG  = "92400E"
    GRIS_HDR  = "F1F5F9"
    GRIS_BORD = "CBD5E1"

    thin = Side(style="thin", color=GRIS_BORD)
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    regimen_label = {"general": "General (30 d/año)", "remype": "REMYPE (15 d/año)"}.get(
        cfg.regimen, f"Otro ({cfg.dias_por_anio} d/año)"
    )

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "REPORTE DE VACACIONES POR LEY"
    c.font = Font(name="Calibri", bold=True, size=14, color=BLANCO)
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:K2")
    c = ws["A2"]
    c.value = (
        f"Régimen: {regimen_label}   |   "
        f"Tope acumulación: {cfg.tope_acumulacion} días   |   "
        f"Generado: {date.today().strftime('%d/%m/%Y')}"
    )
    c.font = Font(name="Calibri", size=10, color="475569")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6   # espacio

    # ── Encabezados ───────────────────────────────────────────────────────────
    COLS = [
        ("N°",              6),
        ("Empleado",        28),
        ("Cargo",           22),
        ("F. Ingreso",      13),
        ("Meses serv.",     13),
        ("Años serv.",      11),
        ("Días/año",        10),
        ("Devengado",       12),
        ("Gozado",          10),
        ("Disponible",      12),
        ("Estado",          22),
    ]
    HDR_ROW = 4
    for ci, (titulo, ancho) in enumerate(COLS, 1):
        col_letter = get_column_letter(ci)
        ws.column_dimensions[col_letter].width = ancho
        cell = ws.cell(row=HDR_ROW, column=ci, value=titulo)
        cell.font = Font(name="Calibri", bold=True, size=10, color="1E293B")
        cell.fill = PatternFill("solid", fgColor=GRIS_HDR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde
    ws.row_dimensions[HDR_ROW].height = 24

    # ── Filas de datos ────────────────────────────────────────────────────────
    for idx, s in enumerate(saldos, 1):
        row = HDR_ROW + idx
        estado = s.estado_vacaciones

        if estado == "sin_derecho":
            bg, fg = AMBAR_BG, AMBAR_FG
        elif estado == "agotado":
            bg, fg = ROJO_BG, ROJO_FG
        else:
            bg, fg = VERDE_BG, VERDE_FG

        valores = [
            idx,
            s.empleado_nombre or "",
            s.cargo or "",
            s.fecha_ingreso or "",
            s.meses_servicio,
            s.anos_servicio,
            s.dias_por_anio,
            s.devengado,
            s.gozado,
            s.disponible,
            _ESTADO_LABEL.get(estado, estado),
        ]
        for ci, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = Font(name="Calibri", size=10,
                             color=(fg if ci == 11 else "1E293B"),
                             bold=(ci == 11))
            if ci == 11:
                cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(
                horizontal="center" if ci in (1, 4, 5, 6, 7, 8, 9, 10) else "left",
                vertical="center"
            )
            cell.border = borde
        ws.row_dimensions[row].height = 20

    # ── Congelar encabezado ───────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)

    # ── Hoja 2: Detalle de períodos aprobados ─────────────────────────────────
    ws2 = wb.create_sheet("Detalle Períodos")
    DET_COLS = [
        ("Empleado",      28), ("Cargo",      18), ("Desde",     13),
        ("Hasta",         13), ("Días",        8), ("Aprobado el", 15),
    ]
    for ci, (titulo, ancho) in enumerate(DET_COLS, 1):
        col_l = get_column_letter(ci)
        ws2.column_dimensions[col_l].width = ancho
        c2 = ws2.cell(row=1, column=ci, value=titulo)
        c2.font = Font(name="Calibri", bold=True, size=10, color="1E293B")
        c2.fill = PatternFill("solid", fgColor="1E3A5F")
        c2.font = Font(name="Calibri", bold=True, size=10, color=BLANCO)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = borde
    ws2.row_dimensions[1].height = 22

    det_row = 2
    for s in saldos:
        if not s.solicitudes_gozadas:
            continue
        for per in s.solicitudes_gozadas:
            vals = [
                s.empleado_nombre or "",
                s.cargo or "",
                per.fecha_inicio,
                per.fecha_fin,
                per.dias,
                per.fecha_aprobacion[:10] if per.fecha_aprobacion else "",
            ]
            for ci, val in enumerate(vals, 1):
                c2 = ws2.cell(row=det_row, column=ci, value=val)
                c2.font = Font(name="Calibri", size=10)
                c2.alignment = Alignment(
                    horizontal="center" if ci > 2 else "left", vertical="center"
                )
                c2.border = borde
                if ci == 5:
                    c2.fill = PatternFill("solid", fgColor=VERDE_BG)
            ws2.row_dimensions[det_row].height = 18
            det_row += 1
    ws2.freeze_panes = ws2.cell(row=2, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"vacaciones_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Reporte PDF ───────────────────────────────────────────────────────────────

def _pdf_footer(canvas, doc, fecha_gen: str, total_pags_ref: list):
    """Pie de página numerado en todas las hojas."""
    from reportlab.lib import colors as _c
    from reportlab.lib.units import cm as _cm
    w, h = doc.pagesize
    canvas.saveState()
    canvas.setFont("Helvetica", 6.5)
    canvas.setFillColor(_c.HexColor("#94A3B8"))
    canvas.drawString(1.5*_cm, 0.7*_cm,
        f"CONFIDENCIAL — Uso exclusivo de Gerencia y Recursos Humanos  |  "
        f"Generado: {fecha_gen}  |  "
        f"Normativa: D.Leg. 713 · D.S. 013-2013-PRODUCE · D.Leg. 1405")
    canvas.drawRightString(w - 1.5*_cm, 0.7*_cm, f"Página {doc.page}")
    canvas.setStrokeColor(_c.HexColor("#E2E8F0"))
    canvas.setLineWidth(0.5)
    canvas.line(1.5*_cm, 1*_cm, w - 1.5*_cm, 1*_cm)
    canvas.restoreState()


@router.get("/reporte.pdf")
def reporte_pdf(payload: dict = Depends(verificar_token), db: Session = Depends(get_db)):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        HRFlowable, KeepTogether, PageBreak,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY

    empresa_id = payload["empresa_id"]
    saldos, cfg = _saldos_todos(db, empresa_id)
    hoy = date.today()
    fecha_gen = hoy.strftime("%d/%m/%Y")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=1.6*cm, bottomMargin=1.8*cm,
    )
    PAGE_W = landscape(A4)[0] - 3.6*cm

    # ── Paleta corporativa ────────────────────────────────────────────────────
    AZUL_OSC  = colors.HexColor("#1E3A5F")
    AZUL_MED  = colors.HexColor("#1D4ED8")
    AZUL_CLR  = colors.HexColor("#EFF6FF")
    GRIS_CLR  = colors.HexColor("#F8FAFC")
    GRIS_BRD  = colors.HexColor("#CBD5E1")
    GRIS_BRD2 = colors.HexColor("#E2E8F0")
    ROJO_BG   = colors.HexColor("#FEF2F2")
    ROJO_FG   = colors.HexColor("#B91C1C")
    VERDE_BG  = colors.HexColor("#F0FDF4")
    VERDE_FG  = colors.HexColor("#15803D")
    AMBAR_BG  = colors.HexColor("#FFFBEB")
    AMBAR_FG  = colors.HexColor("#B45309")
    BLANCO    = colors.white
    NEGRO     = colors.HexColor("#0F172A")
    PLOMO     = colors.HexColor("#475569")

    regimen_label = {
        "general": "Régimen General",
        "remype":  "Régimen REMYPE (Pequeña Empresa)",
    }.get(cfg.regimen, "Régimen especial")

    def _fmt(iso: Optional[str]) -> str:
        if not iso:
            return "—"
        try:
            return date.fromisoformat(iso[:10]).strftime("%d/%m/%Y")
        except Exception:
            return iso

    def PS(name, **kw):
        return ParagraphStyle(name, **kw)

    SEC_ST  = PS("SEC", fontSize=10, textColor=AZUL_OSC, fontName="Helvetica-Bold",
                 spaceBefore=14, spaceAfter=5, leading=13)
    CELL_ST = PS("C",   fontSize=8,  fontName="Helvetica", leading=10, alignment=TA_LEFT)
    CTR_ST  = PS("CT",  fontSize=8,  fontName="Helvetica", leading=10, alignment=TA_CENTER)
    CTR_B   = PS("CTB", fontSize=8,  fontName="Helvetica-Bold", leading=10, alignment=TA_CENTER)
    # Estilos blancos para encabezados sobre fondo oscuro
    CELL_W  = PS("CW",  fontSize=8,  fontName="Helvetica-Bold", leading=10,
                 alignment=TA_LEFT,   textColor=BLANCO)
    CTR_W   = PS("CTW", fontSize=8,  fontName="Helvetica-Bold", leading=10,
                 alignment=TA_CENTER, textColor=BLANCO)
    SML_ST  = PS("SM",  fontSize=7.5, fontName="Helvetica", leading=9.5,
                 alignment=TA_LEFT, textColor=PLOMO)
    SML_CTR = PS("SMC", fontSize=7.5, fontName="Helvetica", leading=9.5,
                 alignment=TA_CENTER, textColor=PLOMO)

    story = []

    # ════════════════════════════════════════════════════════════════════════
    # BANNER — centrado completo
    # ════════════════════════════════════════════════════════════════════════
    banner = Table([
        [Paragraph("REPORTE DE VACACIONES",
                   PS("TT", fontSize=20, fontName="Helvetica-Bold", textColor=BLANCO,
                      alignment=TA_CENTER, leading=24))],
        [Paragraph("Control de Saldos Vacacionales por Ley &nbsp;·&nbsp; Recursos Humanos",
                   PS("TS", fontSize=9, fontName="Helvetica",
                      textColor=colors.HexColor("#BFDBFE"),
                      alignment=TA_CENTER, leading=12))],
        [Paragraph("CONFIDENCIAL &nbsp;·&nbsp; Uso exclusivo de Gerencia y Recursos Humanos",
                   PS("TC", fontSize=7.5, fontName="Helvetica-Bold",
                      textColor=colors.HexColor("#FEF3C7"),
                      alignment=TA_CENTER, leading=11))],
    ], colWidths=[PAGE_W])
    banner.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), AZUL_OSC),
        ("TOPPADDING",    (0, 0), (-1, 0),  14),
        ("TOPPADDING",    (0, 1), (-1, 1),  3),
        ("TOPPADDING",    (0, 2), (-1, 2),  6),
        ("BOTTOMPADDING", (0, 0), (-1, 0),  2),
        ("BOTTOMPADDING", (0, 1), (-1, 1),  2),
        ("BOTTOMPADDING", (0, 2), (-1, 2),  12),
        ("LEFTPADDING",   (0, 0), (-1, -1), 16),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 16),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(banner)

    # Banda de metadatos
    meta = Table([[
        Paragraph(f"<b>Régimen:</b> {regimen_label}", SML_ST),
        Paragraph(f"<b>Días por año:</b> {cfg.dias_por_anio}", SML_ST),
        Paragraph(f"<b>Tope acumulación:</b> {cfg.tope_acumulacion} días", SML_ST),
        Paragraph(f"<b>Fecha de emisión:</b> {fecha_gen}", SML_ST),
        Paragraph(f"<b>Empleados activos:</b> {len(saldos)}", SML_ST),
    ]], colWidths=[PAGE_W / 5] * 5)
    meta.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), AZUL_CLR),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW",     (0, 0), (-1, -1), 0.5, GRIS_BRD),
    ]))
    story.append(meta)
    story.append(Spacer(1, 0.4*cm))

    # ════════════════════════════════════════════════════════════════════════
    # KPIs EJECUTIVOS
    # ════════════════════════════════════════════════════════════════════════
    con_derecho  = [s for s in saldos if s.meses_servicio >= 12]
    total_gozado = sum(s.gozado for s in saldos)
    total_disp   = sum(s.disponible for s in saldos)

    def _kpi_cell(valor, label, sub, bg, fg):
        t = Table([
            [Paragraph(str(valor), PS(f"KV{valor}", fontSize=20, fontName="Helvetica-Bold",
                                      textColor=fg, alignment=TA_CENTER, leading=22))],
            [Paragraph(label,      PS(f"KL{label}", fontSize=8, fontName="Helvetica-Bold",
                                      textColor=fg, alignment=TA_CENTER, leading=10))],
            [Paragraph(sub,        PS(f"KS{sub}",  fontSize=7, fontName="Helvetica",
                                      textColor=fg, alignment=TA_CENTER, leading=9))],
        ], colWidths=[None])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), bg),
            ("TOPPADDING",    (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return t

    kpi_w = PAGE_W / 4
    kpis = Table([[
        _kpi_cell(len(saldos),         "TOTAL EMPLEADOS",  "Activos en el sistema",          AZUL_CLR, AZUL_OSC),
        _kpi_cell(len(con_derecho),    "CON DERECHO",      "Con 1 año o más de servicio",    VERDE_BG, VERDE_FG),
        _kpi_cell(f"{total_gozado}d",  "DÍAS GOZADOS",     "Vacaciones ya tomadas",          AMBAR_BG, AMBAR_FG),
        _kpi_cell(f"{int(total_disp)}d","DÍAS DISPONIBLES","Pendientes de goce",             ROJO_BG,  ROJO_FG),
    ]], colWidths=[kpi_w]*4)
    kpis.setStyle(TableStyle([
        ("LINEAFTER",     (0, 0), (2, -1), 0.5, GRIS_BRD2),
        ("BOX",           (0, 0), (-1, -1), 0.8, GRIS_BRD),
        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(kpis)
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.6, color=GRIS_BRD2, spaceAfter=4))

    # ════════════════════════════════════════════════════════════════════════
    # SECCIÓN 1 — RESUMEN DE SALDOS
    # ════════════════════════════════════════════════════════════════════════
    story.append(Paragraph("1.  Resumen de Saldos por Empleado", SEC_ST))

    hdr_res = [
        Paragraph("N°",         CTR_W),
        Paragraph("Empleado",   CELL_W),
        Paragraph("Cargo",      CELL_W),
        Paragraph("F. Ingreso", CTR_W),
        Paragraph("Meses",      CTR_W),
        Paragraph("Años",       CTR_W),
        Paragraph("Días/Año",   CTR_W),
        Paragraph("Devengado",  CTR_W),
        Paragraph("Gozado",     CTR_W),
        Paragraph("Disponible", CTR_W),
        Paragraph("Estado",     CTR_W),
    ]
    cw_res = [0.8*cm, 5*cm, 4*cm, 2.3*cm, 1.7*cm, 1.5*cm,
              2*cm, 2.1*cm, 1.8*cm, 2.1*cm, None]
    cw_res[-1] = PAGE_W - sum(cw_res[:-1])

    data_res = [hdr_res]
    row_clrs: list = []

    for idx, s in enumerate(saldos, 1):
        est = s.estado_vacaciones
        if est == "sin_derecho":
            row_bg, est_fg, est_txt = AMBAR_BG, AMBAR_FG, "Sin derecho"
        elif est == "agotado":
            row_bg, est_fg, est_txt = ROJO_BG,  ROJO_FG,  "Saldo agotado"
        else:
            row_bg, est_fg, est_txt = VERDE_BG, VERDE_FG, f"{s.disponible:.0f} días disp."
        row_clrs.append(row_bg)

        est_st = PS(f"ES{idx}", fontSize=7.5, fontName="Helvetica-Bold",
                    textColor=est_fg, alignment=TA_CENTER, leading=9)
        data_res.append([
            Paragraph(str(idx),                                    CTR_ST),
            Paragraph(s.empleado_nombre or "",                     CELL_ST),
            Paragraph(s.cargo or "",                               CELL_ST),
            Paragraph(_fmt(s.fecha_ingreso),                       CTR_ST),
            Paragraph(str(s.meses_servicio),                       CTR_ST),
            Paragraph(str(s.anos_servicio),                        CTR_ST),
            Paragraph(str(s.dias_por_anio),                        CTR_ST),
            Paragraph(str(int(s.devengado)),                       CTR_ST),
            Paragraph(f"<b>{s.gozado}</b>" if s.gozado else "0",  CTR_ST),
            Paragraph(f"<b>{s.disponible:.0f}</b>",                CTR_B),
            Paragraph(est_txt,                                     est_st),
        ])

    t_res = Table(data_res, colWidths=cw_res, repeatRows=1)
    ts_res = TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), AZUL_OSC),
        ("TEXTCOLOR",     (0, 0), (-1, 0), BLANCO),
        ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",          (0, 0), (-1, -1), 0.35, GRIS_BRD2),
        ("LINEBELOW",     (0, 0), (-1, 0), 1.5, AZUL_MED),
        ("TOPPADDING",    (0, 0), (-1, -1), 3.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
    ])
    for ri, bg in enumerate(row_clrs, 1):
        ts_res.add("BACKGROUND", (0, ri), (-1, ri), bg)
    t_res.setStyle(ts_res)
    story.append(t_res)

    story.append(Spacer(1, 0.25*cm))
    ley_st = PS("LEY", fontSize=7.5, textColor=PLOMO, alignment=TA_CENTER, leading=11)
    story.append(Paragraph(
        "<font color='#B45309'><b>■</b> Amarillo</font> = Sin derecho: aún no cumple 1 año de servicio"
        " &nbsp;&nbsp; "
        "<font color='#B91C1C'><b>■</b> Rojo</font> = Saldo agotado: ya usó todos sus días devengados"
        " &nbsp;&nbsp; "
        "<font color='#15803D'><b>■</b> Verde</font> = Tiene días disponibles para tomar vacaciones",
        ley_st,
    ))

    # ════════════════════════════════════════════════════════════════════════
    # SECCIÓN 2 — MARCO LEGAL
    # ════════════════════════════════════════════════════════════════════════
    story.append(PageBreak())
    story.append(Paragraph("2.  Marco Legal y Régimen de Vacaciones (SUNAFIL)", SEC_ST))

    # Recuadro de régimen aplicable — centrado
    reg_box = Table([[
        Paragraph(
            f"<b>Régimen aplicado: {regimen_label}</b><br/>"
            f"<font size='8.5'>"
            f"{cfg.dias_por_anio} días de vacaciones por cada año completo de servicio &nbsp;·&nbsp; "
            f"Máximo acumulable: {cfg.tope_acumulacion} días ({cfg.tope_acumulacion // cfg.dias_por_anio} períodos)"
            f"</font>",
            PS("RB", fontSize=9.5, fontName="Helvetica", textColor=AZUL_OSC,
               alignment=TA_CENTER, leading=14),
        )
    ]], colWidths=[PAGE_W])
    reg_box.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), AZUL_CLR),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING",   (0, 0), (-1, -1), 14),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 14),
        ("LINEBELOW",     (0, 0), (-1, -1), 2.5, AZUL_MED),
    ]))
    story.append(reg_box)
    story.append(Spacer(1, 0.35*cm))

    # Tabla de artículos legales — sin guiones, lenguaje claro
    leg_hdr = [
        Paragraph("Norma Legal",        CTR_W),
        Paragraph("Tema",               CTR_W),
        Paragraph("¿Qué significa?",    CELL_W),
        Paragraph("¿Cómo aplica aquí?", CELL_W),
    ]
    leg_cw = [3*cm, 3.8*cm, None, 5.2*cm]
    leg_cw[2] = PAGE_W - sum(c for c in leg_cw if c)

    def _leg_row(i, norma, tema, desc, aplica):
        bg_col = GRIS_CLR if i % 2 == 0 else BLANCO
        return [
            Paragraph(norma, PS(f"LN{i}", fontSize=7.5, fontName="Helvetica-Bold",
                                textColor=AZUL_OSC, alignment=TA_CENTER, leading=11)),
            Paragraph(tema,  PS(f"LT{i}", fontSize=7.5, fontName="Helvetica-Bold",
                                textColor=NEGRO, alignment=TA_CENTER, leading=11)),
            Paragraph(desc,  PS(f"LD{i}", fontSize=7.5, fontName="Helvetica",
                                textColor=PLOMO, leading=11.5, alignment=TA_JUSTIFY)),
            Paragraph(aplica, PS(f"LA{i}", fontSize=7.5, fontName="Helvetica",
                                 textColor=NEGRO, leading=11.5)),
        ]

    leg_rows = [
        _leg_row(1,
            "D. Leg. N° 713\n(Art. 10)",
            "¿Cuándo nace el\nderecho a vacaciones?",
            "El trabajador gana el derecho a vacaciones cuando cumple 1 año completo "
            "de trabajo continuo y ha laborado al menos el 80% de los días hábiles del año. "
            "Antes de eso, no corresponde ningún día de descanso vacacional.",
            "Los empleados que aparecen como 'Sin derecho' en este reporte "
            "todavía no han cumplido su primer año. Su derecho se activa "
            "al completar los 12 meses.",
        ),
        _leg_row(2,
            "D. Leg. N° 713\n(Art. 12)",
            "Días de vacaciones\n(Régimen General)",
            "En el régimen laboral general, cada trabajador tiene derecho a 30 días "
            "de vacaciones pagadas por año de servicio. Este descanso es obligatorio "
            "y debe ser remunerado por la empresa.",
            "No aplica a esta empresa, que se encuentra inscrita en el REMYPE "
            "(régimen de Pequeña Empresa)." if cfg.regimen == "remype"
            else f"Se aplican {cfg.dias_por_anio} días de vacaciones por año completo.",
        ),
        _leg_row(3,
            "D.S. N° 013-2013\nPRODUCE (Art. 49)\nREMYPE",
            "Régimen de Pequeña\nEmpresa (REMYPE)",
            "Las empresas inscritas en el REMYPE otorgan 15 días de vacaciones por año "
            "de servicio, en lugar de los 30 días del régimen general. "
            "Esto es un beneficio legal para las pequeñas empresas que les permite "
            "reducir sus costos laborales manteniéndose dentro de la ley.",
            f"Esta empresa aplica {cfg.dias_por_anio} días de vacaciones por año. "
            f"El máximo que puede acumular un trabajador es {cfg.tope_acumulacion} días.",
        ),
        _leg_row(4,
            "D. Leg. N° 713\n(Arts. 19 y 23)\nAcumulación",
            "¿Se pueden acumular\nlas vacaciones?",
            "Sí, pero con un límite. Las vacaciones pueden acumularse hasta 2 períodos "
            f"consecutivos. Para el REMYPE esto significa un máximo de {cfg.tope_acumulacion} días. "
            "Una vez alcanzado ese tope, los días que se siguen generando se pierden "
            "automáticamente y ya no se acumulan más.",
            f"El tope máximo vigente es de {cfg.tope_acumulacion} días "
            f"({cfg.tope_acumulacion // cfg.dias_por_anio} períodos). "
            "Los trabajadores que llegan a ese límite deben programar sus vacaciones "
            "con urgencia para no perder días.",
        ),
        _leg_row(5,
            "D. Leg. N° 713\n(Art. 23)\nPrescripción",
            "¿Qué pasa si la empresa\nno otorga vacaciones?",
            "Si la empresa no le da vacaciones al trabajador en el plazo que corresponde, "
            "la ley obliga a pagar el triple:\n"
            "El sueldo del mes vacacional.\n"
            "Un sueldo adicional por haber trabajado en ese período.\n"
            "Una indemnización igual a un sueldo mensual.\n"
            "Este derecho a cobrar el triple prescribe a los 2 años.",
            "Las vacaciones no otorgadas generan una deuda laboral (pasivo) para la empresa. "
            "Se recomienda programar el goce antes de que venza el período correspondiente.",
        ),
        _leg_row(6,
            "D. Leg. N° 1405\n(12/09/2018)\nFraccionamiento",
            "¿Se pueden tomar\nen partes?",
            "Sí. El trabajador puede dividir sus vacaciones en partes más pequeñas, "
            "siempre que cada parte sea de al menos 7 días. "
            "Por ejemplo, puede tomar 7 días en marzo y 8 días en agosto. "
            "Esto debe coordinarse previamente con la empresa.",
            "Los trabajadores con días disponibles pueden solicitar vacaciones fraccionadas "
            "coordinando fechas con el área de Recursos Humanos.",
        ),
        _leg_row(7,
            "D. Leg. N° 713\n(Art. 16)\nPago",
            "¿Cuándo se paga\nel sueldo vacacional?",
            "El sueldo del período vacacional debe pagarse ANTES de que el trabajador "
            "empiece su descanso. No se puede pagar después. "
            "El monto es equivalente al sueldo mensual habitual del trabajador.",
            "Coordinar con el área de Planillas para que el pago se realice "
            "con anticipación al inicio de cada período vacacional.",
        ),
    ]

    t_leg = Table([leg_hdr, *leg_rows], colWidths=leg_cw, repeatRows=1)
    t_leg.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), AZUL_OSC),
        ("TEXTCOLOR",     (0, 0), (-1, 0), BLANCO),
        ("ALIGN",         (0, 0), (1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("GRID",          (0, 0), (-1, -1), 0.3, GRIS_BRD2),
        ("LINEBELOW",     (0, 0), (-1, 0), 1.5, AZUL_MED),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 7),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 7),
        ("ROWBACKGROUND", (0, 1), (-1, -1), [BLANCO, GRIS_CLR]),
    ]))
    story.append(t_leg)
    story.append(Spacer(1, 0.4*cm))

    # Cuadro de recomendaciones — cada punto en su propia línea
    REC_TITLE = PS("RH", fontSize=9, fontName="Helvetica-Bold", textColor=AZUL_OSC,
                   alignment=TA_LEFT, leading=12)
    REC_NUM   = PS("RN", fontSize=9, fontName="Helvetica-Bold", textColor=AZUL_MED,
                   alignment=TA_CENTER, leading=12)
    REC_TXT   = PS("RT", fontSize=8.5, fontName="Helvetica", textColor=NEGRO,
                   leading=12.5, alignment=TA_JUSTIFY)

    recomendaciones = [
        ("1", "Programar vacaciones pendientes",
         "Los trabajadores con saldo agotado o que están cerca del tope de 30 días deben "
         "tomar sus vacaciones pronto. Si no lo hacen, la empresa podría estar obligada a "
         "pagar una indemnización equivalente al triple del sueldo mensual."),
        ("2", "Saldo agotado no significa pasivo adicional",
         "Los trabajadores que aparecen con saldo 0 (agotado) ya utilizaron todos sus días "
         "devengados. Esto significa que no representan ninguna deuda adicional para la empresa "
         "durante el año en curso."),
        ("3", "Trabajadores sin derecho",
         "Los empleados marcados como 'Sin derecho' todavía no completan su primer año de servicio. "
         "No corresponde otorgarles vacaciones aún. Su primer período vacacional se activa "
         "automáticamente al cumplir los 12 meses continuos de trabajo."),
        ("4", "Actualizar el registro mensualmente",
         "Se recomienda revisar y actualizar este reporte cada mes para mantener un control "
         "preciso del pasivo laboral por vacaciones y evitar contingencias con SUNAFIL."),
    ]

    rec_rows_data = []
    for num, titulo, texto in recomendaciones:
        rec_rows_data.append([
            Paragraph(num,    REC_NUM),
            Paragraph(f"<b>{titulo}.</b> {texto}", REC_TXT),
        ])

    rec_table = Table(rec_rows_data, colWidths=[0.8*cm, PAGE_W - 0.8*cm])
    rec_table.setStyle(TableStyle([
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (0, -1),  0),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
        ("LINEBELOW",     (0, 0), (-1, -2), 0.4, GRIS_BRD2),
    ]))

    rec_box = Table([
        [Paragraph("Recomendaciones de Gestión para Gerencia", REC_TITLE)],
        [rec_table],
    ], colWidths=[PAGE_W])
    rec_box.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), AZUL_CLR),
        ("TOPPADDING",    (0, 0), (-1, 0),  10),
        ("BOTTOMPADDING", (0, 0), (-1, 0),  4),
        ("TOPPADDING",    (0, 1), (-1, 1),  0),
        ("BOTTOMPADDING", (0, 1), (-1, 1),  6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 14),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 14),
        ("LINERIGHT",     (0, 0), (0, -1),  4, AZUL_MED),
    ]))

    story.append(rec_box)

    # ════════════════════════════════════════════════════════════════════════
    # SECCIÓN 3 — DETALLE POR EMPLEADO (sólo quienes tienen períodos aprobados)
    # ════════════════════════════════════════════════════════════════════════
    con_gozado = [s for s in saldos if s.solicitudes_gozadas]
    if con_gozado:
        story.append(PageBreak())
        story.append(Paragraph("3.  Detalle de Períodos de Vacaciones Gozadas por Empleado", SEC_ST))
        story.append(Paragraph(
            "Se listan únicamente los trabajadores con al menos una solicitud de vacaciones aprobada "
            "y efectivamente gozada. Los días mostrados en 'Gozado' se descuentan del saldo disponible.",
            PS("DET_INT", fontSize=8, textColor=PLOMO, fontName="Helvetica",
               leading=12, spaceAfter=8, alignment=TA_JUSTIFY),
        ))

        for s in con_gozado:
            bloque = []

            # Ficha del empleado
            ficha = Table([[
                Paragraph(f"<b>{s.empleado_nombre or '—'}</b>", CELL_ST),
                Paragraph(s.cargo or "", SML_ST),
                Paragraph(f"Ingreso: {_fmt(s.fecha_ingreso)}", SML_ST),
                Paragraph(f"{s.meses_servicio} meses / {s.anos_servicio} años", SML_ST),
                Paragraph(
                    f"Dev: <b>{int(s.devengado)}</b>d  "
                    f"Gozado: <b>{s.gozado}</b>d  "
                    f"Disp: <b>{s.disponible:.0f}</b>d",
                    PS("FC", fontSize=8, fontName="Helvetica-Bold", textColor=AZUL_OSC,
                       alignment=TA_RIGHT, leading=10),
                ),
            ]], colWidths=[5*cm, 3.5*cm, 2.8*cm, 3*cm, None])
            ficha_cw5 = PAGE_W - 5*cm - 3.5*cm - 2.8*cm - 3*cm
            ficha = Table([[
                Paragraph(f"<b>{s.empleado_nombre or '—'}</b>", CELL_ST),
                Paragraph(s.cargo or "", SML_ST),
                Paragraph(f"Ingreso: {_fmt(s.fecha_ingreso)}", SML_ST),
                Paragraph(f"{s.meses_servicio} meses / {s.anos_servicio} años", SML_ST),
                Paragraph(
                    f"Dev: <b>{int(s.devengado)}</b>d &nbsp; "
                    f"Gozado: <b>{s.gozado}</b>d &nbsp; "
                    f"Disp: <b>{s.disponible:.0f}</b>d",
                    PS(f"FC{s.empleado_id}", fontSize=8, fontName="Helvetica-Bold",
                       textColor=AZUL_OSC, alignment=TA_RIGHT, leading=10),
                ),
            ]], colWidths=[5*cm, 3.5*cm, 2.8*cm, 3*cm, ficha_cw5])
            ficha.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, -1), GRIS_CLR),
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING",   (0, 0), (-1, -1), 8),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
                ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
                ("LINERIGHT",     (0, 0), (0, -1), 3, AZUL_MED),
            ]))
            bloque.append(ficha)

            # Tabla de períodos
            per_hdr = [
                Paragraph("#",              CTR_W),
                Paragraph("Desde",         CTR_W),
                Paragraph("Hasta",         CTR_W),
                Paragraph("Días gozados",  CTR_W),
                Paragraph("Aprobado el",   CTR_W),
                Paragraph("Observación",   CELL_W),
            ]
            per_cw = [0.8*cm, 2.8*cm, 2.8*cm, 2.5*cm, 2.8*cm, None]
            per_cw[-1] = PAGE_W - sum(c for c in per_cw if c)

            per_data = [per_hdr]
            for i, per in enumerate(s.solicitudes_gozadas, 1):
                per_data.append([
                    Paragraph(str(i),                   SML_CTR),
                    Paragraph(_fmt(per.fecha_inicio),   SML_CTR),
                    Paragraph(_fmt(per.fecha_fin),      SML_CTR),
                    Paragraph(f"<b>{per.dias}</b> día{'s' if per.dias != 1 else ''}", SML_CTR),
                    Paragraph(_fmt(per.fecha_aprobacion), SML_CTR),
                    Paragraph("Período aprobado y gozado", SML_ST),
                ])

            t_per = Table(per_data, colWidths=per_cw)
            t_per.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#334155")),
                ("TEXTCOLOR",     (0, 0), (-1, 0), BLANCO),
                ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
                ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
                ("GRID",          (0, 0), (-1, -1), 0.3, GRIS_BRD2),
                ("TOPPADDING",    (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING",   (0, 0), (-1, -1), 5),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
                ("ROWBACKGROUND", (0, 1), (-1, -1), [BLANCO, GRIS_CLR]),
                ("BACKGROUND",    (3, 1), (3, -1), VERDE_BG),
            ]))
            bloque.append(t_per)
            bloque.append(Spacer(1, 0.3*cm))
            story.append(KeepTogether(bloque))

    # ════════════════════════════════════════════════════════════════════════
    # BUILD
    # ════════════════════════════════════════════════════════════════════════
    footer_fn = lambda canvas, doc: _pdf_footer(canvas, doc, fecha_gen, [])
    doc.build(story, onFirstPage=footer_fn, onLaterPages=footer_fn)

    buf.seek(0)
    filename = f"reporte_vacaciones_{hoy.isoformat()}.pdf"
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
